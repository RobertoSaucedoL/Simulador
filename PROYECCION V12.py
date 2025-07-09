import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
import google.generativeai as genai
import json
import os
import numpy as np
from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime
import time

# --- CONFIGURACIÓN DE LA PÁGINA (Debe ser el primer comando de Streamlit) ---
st.set_page_config(
    page_title="Simulador Financiero Jerárquico PORTAWARE",
    layout="wide",
    initial_sidebar_state="expanded"
)

# --- OBTENER RUTA ACTUAL ---
# Usamos una forma robusta de obtener el directorio, que funciona tanto local como en servidores.
current_dir = Path(__file__).parent if "__file__" in locals() else Path.cwd()

# --- RUTA PARA ARCHIVOS DE DATOS (Mejor práctica para organización) ---
DATA_PATH = current_dir / "data"
DATA_PATH.mkdir(exist_ok=True) # Crea la carpeta 'data' si no existe
EXCEL_PATH = DATA_PATH / "PY FINANCIERO V2.xlsx"
SCENARIOS_FILE_PATH = DATA_PATH / "saved_scenarios.json"


# --- OCULTAR ELEMENTOS DE LA INTERFAZ DE STREAMLIT (VERSIÓN SIMPLIFICADA Y ROBUSTA) ---
# Este CSS es más conciso y se enfoca en los selectores principales que Streamlit utiliza.
# Es menos propenso a romperse con futuras actualizaciones.
hide_st_style = """
<style>
#MainMenu {visibility: hidden;}
footer {visibility: hidden;}
header {visibility: hidden;}
.stDeployButton {display: none;}
</style>
"""
st.markdown(hide_st_style, unsafe_allow_html=True)

# --- CONFIGURACIÓN DE GEMINI AI ---
try:
    genai.configure(api_key=st.secrets["GEMINI_API_KEY"])
    GEMINI_AVAILABLE = True
except (FileNotFoundError, KeyError):
    GEMINI_AVAILABLE = False
    st.warning(
        "⚠️ **Advertencia**: La clave de API de Gemini no está configurada en `st.secrets`. "
        "Las funcionalidades de IA no estarán disponibles."
    )

# --- VERIFICACIÓN DEL ARCHIVO EXCEL ---
if not EXCEL_PATH.exists():
    st.error(f"⚠️ Archivo Excel no encontrado en la ruta esperada: {EXCEL_PATH.resolve()}")
    st.info("Por favor, asegúrate de que el archivo 'PY FINANCIERO V2.xlsx' esté dentro de una carpeta llamada 'data' en el mismo directorio que tu script.")
    st.stop()

# --- DEFINICIÓN DE CONSTANTES ---
SHEET_NAME = "PY"
PALETA_GRAFICOS = {
    'Actual': '#1F77B4', 'Simulado': '#FF7F0E', 'Meta': '#2CA02C',
    'Positivo': '#28A745', 'Negativo': '#DC3545', 'Ingresos': '#6F42C1',
    'Egresos': '#FD7E14', 'Neutro': '#6C757D'
}
# (El resto de tus diccionarios de constantes como META_VALUES va aquí sin cambios)
# --- VALORES META HARDCODEADOS ---
META_VALUES = {
    'VENTAS BRUTAS': {
        'VENTAS BRUTAS NACIONAL 16%': {
            'RETAIL': 168934800,
            'CATALOGO': 5232032,
            'MAYOREO': 57000000
        },
        'VENTAS BRUTAS EXTRANJERO': {
            'RETAIL': 0,
            'CATALOGO': 0,
            'MAYOREO': 0
        }
    },
    'DESCUENTOS': 14608084,
    'OTROS INGRESOS': 0,

    'COSTO': {
        'COSTO DIRECTO': {
            'MATERIALES A PROCESO': 101987820,
            'MANO DE OBRA ARMADO': 4859868
        },
        'COSTO INDIRECTO': {
            'COSTOS DE CALIDAD': 159044,
            'COSTOS DE MOLDES': 366000
        },
        'OTROS COSTOS': 0
    },

    'GASTOS_OPERATIVOS_INDIVIDUALES': {
        'SUELDOS Y SALARIOS': 22818917,
        'PRESTACIONES': 0,
        'OTRAS COMPENSACIONES': 0,
        'SEGURIDAD E HIGIENE': 172490,
        'GASTOS DE PERSONAL': 425174,
        'COMBUSTIBLE': 388200,
        'ESTACIONAMIENTO': 147100,
        'TRANSPORTE LOCAL': 180000,
        'GASTOS DE VIAJE': 420000,
        'ASESORIAS PM': 21246,
        'SEGURIDAD Y VIGILANCIA': 41371,
        'SERVICIOS INSTALACIONES': 338864,
        'CELULARES': 144720,
        'SUMINISTROS GENERALES': 144840,
        'SUMINISTROS OFICINA': 66600,
        'SUMINISTROS COMPUTO': 49200,
        'ARRENDAMIENTOS': 6448852,
        'MANTENIMIENTOS': 355000,
        'INVENTARIO FÍSICO': 50000,
        'OTROS IMPUESTOS Y DERECHOS': 13000,
        'NO DEDUCIBLES': 27750,
        'SEGUROS Y FIANZAS': 185033,
        'CAPACITACION Y ENTRENAMIENTO': 131654,
        'MENSAJERIA': 115400,
        'MUESTRAS': 22800,
        'FERIAS Y EXPOSICIONES': 26200,
        'PUBLICIDAD IMPRESA': 67630,
        'IMPRESIONES 3D': 430035,
        'MATERIAL DISEÑO': 18430,
        'PATENTES': 0,
        'LICENCIAS Y SOFTWARE': 470712,
        'ATENCION A CLIENTES': 0,
        'ASESORIAS PF': 725482,
        'PORTALES CLIENTES': 144475,
        'CUOTAS Y SUSCRIPCIONES': 112218,
        'FLETES EXTERNOS': 7594838,
        'FLETES INTERNOS': 0,
        'IMPTOS S/NOMINA': 658364,
        'CONTRIBUCIONES PATRONALES': 3836805,
        'TIMBRES Y FOLIOS FISCALES': 2714,
        'COMISION MERCANTIL': 0,
        'GASTOS ADUANALES': 0
    },
    'TOTAL DE OTROS GASTOS': 0,

    'FINANCIEROS_INDIVIDUALES': {
        'GASTOS FINANCIEROS': 828874.64,
        'PRODUCTOS FINANCIEROS': 52800.00,
        'RESULTADO CAMBIARIO': 981731
    }
}


# --- GESTIÓN DE ESCENARIOS (PERSISTENCIA) ---

def load_scenarios_from_file():
    """
    Carga los escenarios desde el archivo JSON.
    Devuelve un diccionario vacío si el archivo no existe o está corrupto.
    """
    if SCENARIOS_FILE_PATH.exists():
        try:
            with open(SCENARIOS_FILE_PATH, 'r') as f:
                return json.load(f)
        except (json.JSONDecodeError, FileNotFoundError):
            # Si el archivo está corrupto o no se encuentra, se inicia de cero.
            st.error("Error al leer el archivo de escenarios. Se usará una lista vacía.")
            return {}
    return {}

def save_scenarios_to_file():
    """
    Guarda el diccionario de escenarios actual en el archivo JSON.
    IMPORTANTE: Esto solo funcionará de forma persistente si la app se ejecuta
    en un entorno con sistema de archivos persistente (como un servidor local o una VM).
    En plataformas como Streamlit Community Cloud, estos archivos se borrarán.
    """
    try:
        with open(SCENARIOS_FILE_PATH, 'w') as f:
            json.dump(st.session_state.get('saved_scenarios', {}), f, indent=4)
    except Exception as e:
        st.error(f"No se pudo guardar el escenario en el archivo: {e}")

# --- INICIALIZACIÓN DEL ESTADO DE LA SESIÓN ---
def initialize_session_state():
    """
    Inicializa todas las variables necesarias en st.session_state si no existen.
    Esto centraliza la lógica de inicialización.
    """
    if 'app_initialized' not in st.session_state:
        # Carga los escenarios desde el archivo al iniciar la sesión.
        # Esta es la única vez que se leen del disco para evitar sobreescrituras.
        st.session_state['saved_scenarios'] = load_scenarios_from_file()

        # Inicializa las variables de simulación
        estructura = get_cached_structure()
        for cuenta, datos in estructura.items():
            if datos.get('simulable'):
                key = f"sim_{cuenta.replace(' ', '_').replace('/', '_')}"
                st.session_state[key] = 0.0
            if 'subcuentas' in datos:
                for subcuenta, subdatos in datos['subcuentas'].items():
                    if isinstance(subdatos, dict) and 'actual' not in subdatos:
                        for subitem, itemdatos in subdatos.items():
                            if itemdatos.get('simulable'):
                                key = f"sim_{subcuenta.replace(' ', '_')}_{subitem.replace(' ', '_')}"
                                st.session_state[key] = 0.0
                    elif isinstance(subdatos, dict) and subdatos.get('simulable'):
                        key = f"sim_{subcuenta.replace(' ', '_')}"
                        st.session_state[key] = 0.0

        # Inicializa otras variables de control
        st.session_state['ajuste_activo'] = False
        st.session_state['porcentaje_ajuste'] = 45
        st.session_state['app_initialized'] = True


# --- FUNCIONES DE CÁLCULO Y LÓGICA DE LA APP (Tus funciones sin cambios) ---
# ... (Aquí van tus funciones: obtener_valor_celda, cargar_excel, obtener_estructura_cuentas,
# get_cached_structure, get_actual_value, get_meta_value, calculate_account_value,
# generar_dataframe_completo, obtener_variables_modificadas, aplicar_estilo_financiero,
# generar_recomendacion_variables_ia, generar_insight_financiero, etc.)
# ... (Estas funciones parecen correctas, así que las omito por brevedad)
# --- FUNCIÓN PARA OBTENER VALOR DE CELDA (CORRECCIÓN: DEFINICIÓN AGREGADA/MOVIDA AQUÍ) ---
def obtener_valor_celda(cell_address):
    """
    Obtiene el valor de una celda específica del archivo Excel.
    Asume que el formato de la celda es 'ColumnaLetraFilaNumero' (ej. 'I2').
    """
    try:
        wb = load_workbook(EXCEL_PATH, data_only=True)
        sheet = wb[SHEET_NAME]
        cell_value = sheet[cell_address].value
        # Intentar convertir a float. Si el valor es None o no numérico, retorna 0.0
        try:
            return float(cell_value)
        except (ValueError, TypeError):
            return 0.0
    except Exception as e:
        st.warning(f"Error al leer la celda {cell_address} del Excel: {e}. Retornando 0.0.")
        return 0.0

# --- FUNCIÓN PARA CARGAR EXCEL ---
def cargar_excel():
    """Carga el archivo Excel y lo almacena en session_state"""
    try:
        wb = load_workbook(EXCEL_PATH, data_only=True)
        
        # Verificar si la hoja existe
        if SHEET_NAME not in wb.sheetnames:
            st.error(f"Hoja '{SHEET_NAME}' no encontrada en el archivo")
            return False
        
        sheet = wb[SHEET_NAME]
        
        # Convertir hoja a DataFrame (considerando que la primera fila son encabezados)
        data = sheet.values
        cols = next(data) # Obtener la primera fila como encabezados
        df = pd.DataFrame(data, columns=cols)
        
        # Almacenar en session_state
        st.session_state.excel_data = df
        st.session_state.last_modified = datetime.now()
        return True
    except Exception as e:
        st.error(f"Error al cargar Excel: {str(e)}")
        return False

def obtener_estructura_cuentas():
    """Retorna la estructura completa de cuentas con jerarquía para PORTAWARE."""
    # Mapeo de celdas para gastos operativos
    gastos_operativos_map_cells = {
        'SUELDOS Y SALARIOS': 'I20', 'PRESTACIONES': 'I21', 'OTRAS COMPENSACIONES': 'I22',
        'SEGURIDAD E HIGIENE': 'I23', 'GASTOS DE PERSONAL': 'I24', 'COMBUSTIBLE': 'I25',
        'ESTACIONAMIENTO': 'I26', 'TRANSPORTE LOCAL': 'I27', 'GASTOS DE VIAJE': 'I28',
        'ASESORIAS PM': 'I29', 'SEGURIDAD Y VIGILANCIA': 'I30', 'SERVICIOS INSTALACIONES': 'I31',
        'CELULARES': 'I32', 'SUMINISTROS GENERALES': 'I33', 'SUMINISTROS OFICINA': 'I34',
        'SUMINISTROS COMPUTO': 'I35', 'ARRENDAMIENTOS': 'I36', 'MANTENIMIENTOS': 'I37',
        'INVENTARIO FÍSICO': 'I38', 'OTROS IMPUESTOS Y DERECHOS': 'I39', 'NO DEDUCIBLES': 'I40',
        'SEGUROS Y FIANZAS': 'I41', 'CAPACITACION Y ENTRENAMIENTO': 'I42', 'MENSAJERIA': 'I43',
        'MUESTRAS': 'I44', 'FERIAS Y EXPOSICIONES': 'I45', 'PUBLICIDAD IMPRESA': 'I46',
        'IMPRESIONES 3D': 'I47', 'MATERIAL DISEÑO': 'I48', 'PATENTES': 'I49',
        'LICENCIAS Y SOFTWARE': 'I50', 'ATENCION A CLIENTES': 'I51', 'ASESORIAS PF': 'I52',
        'PORTALES CLIENTES': 'I53', 'CUOTAS Y SUSCRIPCIONES': 'I54', 'FLETES EXTERNOS': 'I55',
        'FLETES INTERNOS': 'I56', 'IMPTOS S/NOMINA': 'I57', 'CONTRIBUCIONES PATRONALES': 'I58',
        'TIMBRES Y FOLIOS FISCALES': 'I59', 'COMISION MERCANTIL': 'I60', 'GASTOS ADUANALES': 'I61'
    }

    gastos_operativos_subcuentas = {}
    for gasto, cell in gastos_operativos_map_cells.items():
        gastos_operativos_subcuentas[gasto] = {
            'actual': obtener_valor_celda(cell),
            'meta': META_VALUES['GASTOS_OPERATIVOS_INDIVIDUALES'].get(gasto, 0),
            'simulable': True
        }

    return {
        # Nivel 4 - VENTAS BRUTAS (INGRESO)
        'VENTAS BRUTAS': {
            'jerarquia': '4', 'tipo': 'suma',
            'componentes': ['VENTAS BRUTAS NACIONAL 16%', 'VENTAS BRUTAS EXTRANJERO'],
            'subcuentas': {
                'VENTAS BRUTAS NACIONAL 16%': {
                    'RETAIL': {'actual': obtener_valor_celda('I3'),
                               'meta': META_VALUES['VENTAS BRUTAS']['VENTAS BRUTAS NACIONAL 16%']['RETAIL'],
                               'simulable': True},
                    'CATALOGO': {'actual': obtener_valor_celda('I4'),
                                 'meta': META_VALUES['VENTAS BRUTAS']['VENTAS BRUTAS NACIONAL 16%']['CATALOGO'],
                                 'simulable': True},
                    'MAYOREO': {'actual': obtener_valor_celda('I5'),
                                'meta': META_VALUES['VENTAS BRUTAS']['VENTAS BRUTAS NACIONAL 16%']['MAYOREO'],
                                'simulable': True}
                },
                'VENTAS BRUTAS EXTRANJERO': {
                    'RETAIL': {'actual': obtener_valor_celda('I6'),
                               'meta': META_VALUES['VENTAS BRUTAS']['VENTAS BRUTAS EXTRANJERO']['RETAIL'],
                               'simulable': True},
                    'CATALOGO': {'actual': obtener_valor_celda('I7'),
                                 'meta': META_VALUES['VENTAS BRUTAS']['VENTAS BRUTAS EXTRANJERO']['CATALOGO'],
                                 'simulable': True},
                    'MAYOREO': {'actual': obtener_valor_celda('I8'),
                                'meta': META_VALUES['VENTAS BRUTAS']['VENTAS BRUTAS EXTRANJERO']['MAYOREO'],
                                'simulable': True}
                }
            }
        },

        # Nivel 5 - DESCUENTOS Y OTROS (DESCUENTOS son una reducción de ingresos, OTROS INGRESOS es un ingreso)
        'DESCUENTOS': {'jerarquia': '5', 'tipo': 'simple', 'actual': obtener_valor_celda('I9'),
                       'meta': META_VALUES['DESCUENTOS'],
                       'simulable': True},
        'OTROS INGRESOS': {'jerarquia': '5.3', 'tipo': 'simple', 'actual': obtener_valor_celda('I10'),
                             'meta': META_VALUES['OTROS INGRESOS'],
                             'simulable': True},

        # Nivel 6 - VENTAS NETAS (INGRESO CLAVE)
        'VENTAS NETAS': {'jerarquia': '6', 'tipo': 'formula', 'formula': 'VENTAS BRUTAS - DESCUENTOS + OTROS INGRESOS'},

        # Nivel 7 - COSTOS (EGRESO)
        'COSTO': {
            'jerarquia': '7', 'tipo': 'suma',
            'componentes': ['COSTO DIRECTO', 'COSTO INDIRECTO', 'OTROS COSTOS'],
            'subcuentas': {
                'COSTO DIRECTO': {
                    'MATERIALES A PROCESO': {'actual': obtener_valor_celda('I13'),
                                             'meta': META_VALUES['COSTO']['COSTO DIRECTO']['MATERIALES A PROCESO'],
                                             'simulable': True},
                    'MANO DE OBRA ARMADO': {'actual': obtener_valor_celda('I14'),
                                            'meta': META_VALUES['COSTO']['COSTO DIRECTO']['MANO DE OBRA ARMADO'],
                                            'simulable': True}
                },
                'COSTO INDIRECTO': {
                    'COSTOS DE CALIDAD': {'actual': obtener_valor_celda('I15'),
                                          'meta': META_VALUES['COSTO']['COSTO INDIRECTO']['COSTOS DE CALIDAD'],
                                          'simulable': True},
                    'COSTOS DE MOLDES': {'actual': obtener_valor_celda('I16'),
                                         'meta': META_VALUES['COSTO']['COSTO INDIRECTO']['COSTOS DE MOLDES'],
                                         'simulable': True}
                },
                'OTROS COSTOS': {
                    'OTROS COSTOS': {'actual': obtener_valor_celda('I17'), 'meta': META_VALUES['COSTO']['OTROS COSTOS'],
                                     'simulable': True}}
            }
        },

        # Nivel 8 - MARGEN BRUTO
        'MARGEN BRUTO': {'jerarquia': '8', 'tipo': 'formula', 'formula': 'VENTAS NETAS - COSTO'},

        # Nivel 9 - GASTOS OPERATIVOS (EGRESO)
        'TOTAL GASTOS OPERATIVOS': {
            'jerarquia': '9', 'tipo': 'suma_gastos',
            'subcuentas': gastos_operativos_subcuentas  # Se asigna el diccionario generado dinámicamente
        },

        # Resto de la estructura
        'EBITDA OPERATIVA': {'jerarquia': '10', 'tipo': 'formula',
                             'formula': 'VENTAS NETAS - COSTO - TOTAL GASTOS OPERATIVOS'},
        'TOTAL DE OTROS GASTOS': {'jerarquia': '11', 'tipo': 'simple', 'actual': obtener_valor_celda('I63'),
                                  'meta': META_VALUES['TOTAL DE OTROS GASTOS'],
                                  'simulable': True},
        'EBITDA': {'jerarquia': '12', 'tipo': 'formula', 'formula': 'EBITDA OPERATIVA - TOTAL DE OTROS GASTOS'},
        'FINANCIEROS': {
            'jerarquia': '13', 'tipo': 'suma',
            'componentes': ['GASTOS FINANCIEROS', 'PRODUCTOS FINANCIEROS', 'RESULTADO CAMBIARIO'],
            'subcuentas': {
                # Valores tomados directamente de la imagen para 'meta' y 'actual'
                'GASTOS FINANCIEROS': {'actual': obtener_valor_celda('I66'),
                                       'meta': META_VALUES['FINANCIEROS_INDIVIDUALES']['GASTOS FINANCIEROS'],
                                       'simulable': True},
                'PRODUCTOS FINANCIEROS': {'actual': obtener_valor_celda('I67'),
                                          'meta': META_VALUES['FINANCIEROS_INDIVIDUALES']['PRODUCTOS FINANCIEROS'],
                                          'simulable': True},
                'RESULTADO CAMBIARIO': {'actual': obtener_valor_celda('I68'),
                                        'meta': META_VALUES['FINANCIEROS_INDIVIDUALES']['RESULTADO CAMBIARIO'],
                                        'simulable': True}
            }
        },
        'BAI': {'jerarquia': '14', 'tipo': 'formula', 'formula': 'EBITDA - FINANCIEROS'}
    }

# --- FUNCIONES DE CÁLCULO ---
@st.cache_data
def get_cached_structure():
    return obtener_estructura_cuentas()

def get_actual_value(estructura, account_key, sub_account_key=None, sub_item_key=None):
    if sub_account_key and sub_item_key:
        return estructura.get(account_key, {}).get('subcuentas', {}).get(sub_account_key, {}).get(sub_item_key, {}).get(
            'actual', 0)
    elif sub_account_key:
        return estructura.get(account_key, {}).get('subcuentas', {}).get(sub_account_key, {}).get('actual', 0)
    else:
        return estructura.get(account_key, {}).get('actual', 0)

def get_meta_value(estructura, account_key, sub_account_key=None, sub_item_key=None):
    if sub_account_key and sub_item_key:
        return estructura.get(account_key, {}).get('subcuentas', {}).get(sub_account_key, {}).get(sub_item_key, {}).get(
            'meta', 0)
    elif sub_account_key:
        return estructura.get(account_key, {}).get('subcuentas', {}).get(sub_account_key, {}).get('meta', 0)
    else:
        return estructura.get(account_key, {}).get('meta', 0)

def calculate_account_value(_estructura, scenario, changes):
    results = {}

    def get_val(cuenta):
        if cuenta in results:
            return results[cuenta]

        datos = _estructura.get(cuenta)
        if not datos:
            return 0

        valor = 0
        if datos['tipo'] == 'simple':
            base_value = datos['actual'] if scenario == 'actual' else datos.get('meta', 0) if scenario == 'meta' else \
                datos['actual']
            if datos.get('simulable') and scenario == 'simulado':
                key = f"sim_{cuenta.replace(' ', '_').replace('/', '_')}"
                cambio_monetario = changes.get(key, 0.0)
                valor = base_value + cambio_monetario
            else:
                valor = base_value
        elif datos['tipo'] in ['suma', 'suma_gastos']:
            for subcuenta_name, subdatos_dict_or_item in datos['subcuentas'].items():
                if isinstance(subdatos_dict_or_item, dict) and 'actual' not in subdatos_dict_or_item:
                    for subitem_name, itemdatos in subdatos_dict_or_item.items():
                        base_value = itemdatos['actual'] if scenario == 'actual' else itemdatos.get('meta',
                                                                                                    0) if scenario == 'meta' else \
                            itemdatos['actual']
                        if itemdatos.get('simulable') and scenario == 'simulado':
                            key = f"sim_{subcuenta_name.replace(' ', '_')}_{subitem_name.replace(' ', '_')}"
                            cambio_monetario = changes.get(key, 0.0)
                            valor += base_value + cambio_monetario
                        else:
                            valor += base_value
                else:
                    itemdatos = subdatos_dict_or_item
                    base_value = itemdatos['actual'] if scenario == 'actual' else itemdatos.get('meta',
                                                                                                0) if scenario == 'meta' else \
                        itemdatos['actual']
                    if itemdatos.get('simulable') and scenario == 'simulado':
                        key = f"sim_{subcuenta_name.replace(' ', '_')}"
                        cambio_monetario = changes.get(key, 0.0)
                        valor += base_value + cambio_monetario
                    else:
                        valor += base_value
        elif datos['tipo'] == 'formula':
            temp_formula = datos['formula'].replace(' - ', '|').replace(' + ', '|')
            operands = temp_formula.split('|')
            operators = [char for char in datos['formula'] if char in ['+', '-']]

            first_operand_name = operands[0].strip()
            first_operand_val = results.get(first_operand_name)
            if first_operand_val is None:
                first_operand_val = get_val(first_operand_name)

            valor = first_operand_val
            for i, operand_name_raw in enumerate(operands[1:]):
                op = operators[i]
                current_operand_name = operand_name_raw.strip()
                current_operand_val = results.get(current_operand_name)
                if current_operand_val is None:
                    current_operand_val = get_val(current_operand_name)

                if op == '+':
                    valor += current_operand_val
                elif op == '-':
                    valor -= current_operand_val

        results[cuenta] = valor
        return valor

    cuentas_ordenadas = sorted(_estructura.keys(), key=lambda k: float(_estructura[k]['jerarquia']))

    for _ in range(5):
        for cuenta in cuentas_ordenadas:
            if cuenta not in results or _estructura[cuenta]['tipo'] == 'formula':
                get_val(cuenta)

    for cuenta in cuentas_ordenadas:
        if cuenta not in results:
            if _estructura[cuenta]['tipo'] == 'simple':
                results[cuenta] = _estructura[cuenta]['actual'] if scenario == 'actual' else _estructura[cuenta].get(
                    'meta', 0)
            elif _estructura[cuenta]['tipo'] in ['suma', 'suma_gastos']:
                sum_val = 0
                for subcuenta_name, subdatos_dict_or_item in _estructura[cuenta]['subcuentas'].items():
                    if isinstance(subdatos_dict_or_item, dict) and 'actual' not in subdatos_dict_or_item:
                        for subitem_name, itemdatos in subdatos_dict_or_item.items():
                            sum_val += itemdatos['actual'] if scenario == 'actual' else itemdatos.get('meta', 0)
                    else:
                        itemdatos = subdatos_dict_or_item
                        sum_val += itemdatos['actual'] if scenario == 'actual' else itemdatos.get('meta', 0)
                results[cuenta] = sum_val
    return results

def generar_dataframe_completo(changes):
    estructura = get_cached_structure()
    actual_values = calculate_account_value(estructura, 'actual', changes)
    meta_values = calculate_account_value(estructura, 'meta', changes)
    simulado_values = calculate_account_value(estructura, 'simulado', changes)

    data_list = []
    for c, i in estructura.items():
        data_list.append({
            'Cuenta': c,
            'Jerarquia': i['jerarquia'],
            'Actual': actual_values.get(c, 0),
            'Simulado': simulado_values.get(c, 0),
            'Meta': meta_values.get(c, 0)
        })
    df = pd.DataFrame(data_list)

    df['Jerarquia'] = pd.to_numeric(df['Jerarquia'])
    df = df.sort_values(by='Jerarquia').reset_index(drop=True)

    df['Brecha vs Meta (%)'] = ((df['Simulado'] - df['Meta']) / df['Meta'].replace(0, pd.NA)) * 100
    df['Brecha Simulado vs Actual (%)'] = ((df['Simulado'] - df['Actual']) / df['Actual'].replace(0, pd.NA)) * 100

    # --- LÓGICA PARA NUEVA COLUMNA ---
    ventas_netas_actual = df.loc[df['Cuenta'] == 'VENTAS NETAS', 'Actual'].iloc[0] if 'VENTAS NETAS' in df[
        'Cuenta'].values else 1
    ventas_netas_simulado = df.loc[df['Cuenta'] == 'VENTAS NETAS', 'Simulado'].iloc[0] if 'VENTAS NETAS' in df[
        'Cuenta'].values else 1
    ventas_netas_meta = df.loc[df['Cuenta'] == 'VENTAS NETAS', 'Meta'].iloc[0] if 'VENTAS NETAS' in df[
        'Cuenta'].values else 1
    
    df['Actual (% VN)'] = (df['Actual'] / ventas_netas_actual) * 100 if ventas_netas_actual != 0 else 0
    df['Simulado (% VN)'] = (df['Simulado'] / ventas_netas_simulado) * 100 if ventas_netas_simulado != 0 else 0
    df['Meta (% VN)'] = (df['Meta'] / ventas_netas_meta) * 100 if ventas_netas_meta != 0 else 0

    return df

def obtener_variables_modificadas(changes):
    variables_modificadas = []

    estructura = get_cached_structure()
    sim_key_info = {}
    for cuenta, datos in estructura.items():
        if datos.get('simulable') and 'subcuentas' not in datos:
            sim_key_info[f"sim_{cuenta.replace(' ', '_').replace('/', '_')}"] = {'name': cuenta,
                                                                                 'actual_val': datos['actual']}
        if 'subcuentas' in datos and isinstance(datos['subcuentas'], dict):
            for subcuenta_name, subdatos_value in datos['subcuentas'].items():
                if isinstance(subdatos_value, dict) and 'actual' in subdatos_value and subdatos_value.get('simulable'):
                    full_name = subcuenta_name
                    actual_val = subdatos_value['actual']
                    meta_val = subdatos_value['meta']
                    sim_key_info[f"sim_{subcuenta_name.replace(' ', '_')}"] = {
                        'name': full_name, 'actual_val': subdatos_value['actual']}
                elif isinstance(subdatos_value, dict) and 'actual' not in subdatos_value:
                    for subitem_name, itemdatos in subdatos_value.items():
                        if itemdatos.get('simulable'):
                            full_name = f"{subcuenta_name} - {subitem_name}"
                            actual_val = itemdatos['actual']
                  _Borrador_de_mensaje (editado)
<div class="message-body" data-message-id="2044238545806653112" data-message-author-id="271318042974912513">
```python
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
import google.generativeai as genai
import json
import os
import numpy as np
from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime
import time

# --- CONFIGURACIÓN DE LA PÁGINA (Debe ser el primer comando de Streamlit) ---
st.set_page_config(
    page_title="Simulador Financiero Jerárquico PORTAWARE",
    layout="wide",
    initial_sidebar_state="expanded"
)

# --- OBTENER RUTA ACTUAL ---
# Usamos una forma robusta de obtener el directorio, que funciona tanto local como en servidores.
current_dir = Path(__file__).parent if "__file__" in locals() else Path.cwd()

# --- RUTA PARA ARCHIVOS DE DATOS (Mejor práctica para organización) ---
DATA_PATH = current_dir / "data"
DATA_PATH.mkdir(exist_ok=True) # Crea la carpeta 'data' si no existe
EXCEL_PATH = DATA_PATH / "PY FINANCIERO V2.xlsx"
SCENARIOS_FILE_PATH = DATA_PATH / "saved_scenarios.json"


# --- OCULTAR ELEMENTOS DE LA INTERFAZ DE STREAMLIT (VERSIÓN SIMPLIFICADA Y ROBUSTA) ---
# Este CSS es más conciso y se enfoca en los selectores principales que Streamlit utiliza.
# Es menos propenso a romperse con futuras actualizaciones.
hide_st_style = """
<style>
#MainMenu {visibility: hidden;}
footer {visibility: hidden;}
header {visibility: hidden;}
.stDeployButton {display: none;}
</style>
"""
st.markdown(hide_st_style, unsafe_allow_html=True)

# --- CONFIGURACIÓN DE GEMINI AI ---
try:
    genai.configure(api_key=st.secrets["GEMINI_API_KEY"])
    GEMINI_AVAILABLE = True
except (FileNotFoundError, KeyError):
    GEMINI_AVAILABLE = False
    st.warning(
        "⚠️ **Advertencia**: La clave de API de Gemini no está configurada en `st.secrets`. "
        "Las funcionalidades de IA no estarán disponibles."
    )

# --- VERIFICACIÓN DEL ARCHIVO EXCEL ---
if not EXCEL_PATH.exists():
    st.error(f"⚠️ Archivo Excel no encontrado en la ruta esperada: {EXCEL_PATH.resolve()}")
    st.info("Por favor, asegúrate de que el archivo 'PY FINANCIERO V2.xlsx' esté dentro de una carpeta llamada 'data' en el mismo directorio que tu script.")
    st.stop()

# --- DEFINICIÓN DE CONSTANTES ---
SHEET_NAME = "PY"
PALETA_GRAFICOS = {
    'Actual': '#1F77B4', 'Simulado': '#FF7F0E', 'Meta': '#2CA02C',
    'Positivo': '#28A745', 'Negativo': '#DC3545', 'Ingresos': '#6F42C1',
    'Egresos': '#FD7E14', 'Neutro': '#6C757D'
}
# (El resto de tus diccionarios de constantes como META_VALUES va aquí sin cambios)
# --- VALORES META HARDCODEADOS ---
META_VALUES = {
    'VENTAS BRUTAS': {
        'VENTAS BRUTAS NACIONAL 16%': {
            'RETAIL': 168934800,
            'CATALOGO': 5232032,
            'MAYOREO': 57000000
        },
        'VENTAS BRUTAS EXTRANJERO': {
            'RETAIL': 0,
            'CATALOGO': 0,
            'MAYOREO': 0
        }
    },
    'DESCUENTOS': 14608084,
    'OTROS INGRESOS': 0,

    'COSTO': {
        'COSTO DIRECTO': {
            'MATERIALES A PROCESO': 101987820,
            'MANO DE OBRA ARMADO': 4859868
        },
        'COSTO INDIRECTO': {
            'COSTOS DE CALIDAD': 159044,
            'COSTOS DE MOLDES': 366000
        },
        'OTROS COSTOS': 0
    },

    'GASTOS_OPERATIVOS_INDIVIDUALES': {
        'SUELDOS Y SALARIOS': 22818917,
        'PRESTACIONES': 0,
        'OTRAS COMPENSACIONES': 0,
        'SEGURIDAD E HIGIENE': 172490,
        'GASTOS DE PERSONAL':  425174,
        'COMBUSTIBLE':   388200,
        'ESTACIONAMIENTO':   147100,
        'TRANSPORTE LOCAL': 180000,
        'GASTOS DE VIAJE':  420000,
        'ASESORIAS PM': 21246,
        'SEGURIDAD Y VIGILANCIA': 41371,
        'SERVICIOS INSTALACIONES': 338864,
        'CELULARES': 144720,
        'SUMINISTROS GENERALES': 144840,
        'SUMINISTROS OFICINA': 66600,
        'SUMINISTROS COMPUTO': 49200,
        'ARRENDAMIENTOS': 6448852,
        'MANTENIMIENTOS': 355000,
        'INVENTARIO FÍSICO': 50000,
        'OTROS IMPUESTOS Y DERECHOS':  13000,
        'NO DEDUCIBLES':  27750,
        'SEGUROS Y FIANZAS': 185033,
        'CAPACITACION Y ENTRENAMIENTO': 131654,
        'MENSAJERIA': 115400,
        'MUESTRAS': 22800,
        'FERIAS Y EXPOSICIONES': 26200,
        'PUBLICIDAD IMPRESA': 67630,
        'IMPRESIONES 3D': 430035,
        'MATERIAL DISEÑO': 18430,
        'PATENTES': 0,
        'LICENCIAS Y SOFTWARE': 470712,
        'ATENCION A CLIENTES': 0,
        'ASESORIAS PF':  725482,
        'PORTALES CLIENTES': 144475,
        'CUOTAS Y SUSCRIPCIONES':  112218,
        'FLETES EXTERNOS': 7594838,
        'FLETES INTERNOS': 0,
        'IMPTOS S/NOMINA': 658364,
        'CONTRIBUCIONES PATRONALES': 3836805,
        'TIMBRES Y FOLIOS FISCALES': 2714,
        'COMISION MERCANTIL': 0,
        'GASTOS ADUANALES': 0
    },
    'TOTAL DE OTROS GASTOS': 0,

    'FINANCIEROS_INDIVIDUALES': {
        'GASTOS FINANCIEROS': 828874.64,
        'PRODUCTOS FINANCIEROS': 52800.00,
        'RESULTADO CAMBIARIO':   981731
    }
}


# --- GESTIÓN DE ESCENARIOS (PERSISTENCIA) ---

def load_scenarios_from_file():
    """
    Carga los escenarios desde el archivo JSON.
    Devuelve un diccionario vacío si el archivo no existe o está corrupto.
    """
    if SCENARIOS_FILE_PATH.exists():
        try:
            with open(SCENARIOS_FILE_PATH, 'r') as f:
                return json.load(f)
        except (json.JSONDecodeError, FileNotFoundError):
            # Si el archivo está corrupto o no se encuentra, se inicia de cero.
            st.error("Error al leer el archivo de escenarios. Se usará una lista vacía.")
            return {}
    return {}

def save_scenarios_to_file():
    """
    Guarda el diccionario de escenarios actual en el archivo JSON.
    IMPORTANTE: Esto solo funcionará de forma persistente si la app se ejecuta
    en un entorno con sistema de archivos persistente (como un servidor local o una VM).
    En plataformas como Streamlit Community Cloud, estos archivos se borrarán.
    """
    try:
        with open(SCENARIOS_FILE_PATH, 'w') as f:
            json.dump(st.session_state.get('saved_scenarios', {}), f, indent=4)
    except Exception as e:
        st.error(f"No se pudo guardar el escenario en el archivo: {e}")

# --- INICIALIZACIÓN DEL ESTADO DE LA SESIÓN ---
def initialize_session_state():
    """
    Inicializa todas las variables necesarias en st.session_state si no existen.
    Esto centraliza la lógica de inicialización.
    """
    if 'app_initialized' not in st.session_state:
        # Carga los escenarios desde el archivo al iniciar la sesión.
        # Esta es la única vez que se leen del disco para evitar sobreescrituras.
        st.session_state['saved_scenarios'] = load_scenarios_from_file()

        # Inicializa las variables de simulación
        estructura = get_cached_structure()
        for cuenta, datos in estructura.items():
            if datos.get('simulable'):
                key = f"sim_{cuenta.replace(' ', '_').replace('/', '_')}"
                st.session_state[key] = 0.0
            if 'subcuentas' in datos:
                for subcuenta, subdatos in datos['subcuentas'].items():
                    if isinstance(subdatos, dict) and 'actual' not in subdatos:
                        for subitem, itemdatos in subdatos.items():
                            if itemdatos.get('simulable'):
                                key = f"sim_{subcuenta.replace(' ', '_')}_{subitem.replace(' ', '_')}"
                                st.session_state[key] = 0.0
                    elif isinstance(subdatos, dict) and subdatos.get('simulable'):
                        key = f"sim_{subcuenta.replace(' ', '_')}"
                        st.session_state[key] = 0.0

        # Inicializa otras variables de control
        st.session_state['ajuste_activo'] = False
        st.session_state['porcentaje_ajuste'] = 45
        st.session_state['app_initialized'] = True


# --- FUNCIONES DE CÁLCULO Y LÓGICA DE LA APP (Tus funciones sin cambios) ---
# ... (Aquí van tus funciones: obtener_valor_celda, cargar_excel, obtener_estructura_cuentas,
# get_cached_structure, get_actual_value, get_meta_value, calculate_account_value,
# generar_dataframe_completo, obtener_variables_modificadas, aplicar_estilo_financiero,
# generar_recomendacion_variables_ia, generar_insight_financiero, etc.)
# ... (Estas funciones parecen correctas, así que las omito por brevedad)
# --- FUNCIÓN PARA OBTENER VALOR DE CELDA (CORRECCIÓN: DEFINICIÓN AGREGADA/MOVIDA AQUÍ) ---
def obtener_valor_celda(cell_address):
    """
    Obtiene el valor de una celda específica del archivo Excel.
    Asume que el formato de la celda es 'ColumnaLetraFilaNumero' (ej. 'I2').
    """
    try:
        wb = load_workbook(EXCEL_PATH, data_only=True)
        sheet = wb[SHEET_NAME]
        cell_value = sheet[cell_address].value
        # Intentar convertir a float. Si el valor es None o no numérico, retorna 0.0
        try:
            return float(cell_value)
        except (ValueError, TypeError):
            return 0.0
    except Exception as e:
        st.warning(f"Error al leer la celda {cell_address} del Excel: {e}. Retornando 0.0.")
        return 0.0

# --- FUNCIÓN PARA CARGAR EXCEL ---
def cargar_excel():
    """Carga el archivo Excel y lo almacena en session_state"""
    try:
        wb = load_workbook(EXCEL_PATH, data_only=True)
        
        # Verificar si la hoja existe
        if SHEET_NAME not in wb.sheetnames:
            st.error(f"Hoja '{SHEET_NAME}' no encontrada en el archivo")
            return False
        
        sheet = wb[SHEET_NAME]
        
        # Convertir hoja a DataFrame (considerando que la primera fila son encabezados)
        data = sheet.values
        cols = next(data) # Obtener la primera fila como encabezados
        df = pd.DataFrame(data, columns=cols)
        
        # Almacenar en session_state
        st.session_state.excel_data = df
        st.session_state.last_modified = datetime.now()
        return True
    except Exception as e:
        st.error(f"Error al cargar Excel: {str(e)}")
        return False

def obtener_estructura_cuentas():
    """Retorna la estructura completa de cuentas con jerarquía para PORTAWARE."""
    # Mapeo de celdas para gastos operativos
    gastos_operativos_map_cells = {
        'SUELDOS Y SALARIOS': 'I20', 'PRESTACIONES': 'I21', 'OTRAS COMPENSACIONES': 'I22',
        'SEGURIDAD E HIGIENE': 'I23', 'GASTOS DE PERSONAL': 'I24', 'COMBUSTIBLE': 'I25',
        'ESTACIONAMIENTO': 'I26', 'TRANSPORTE LOCAL': 'I27', 'GASTOS DE VIAJE': 'I28',
        'ASESORIAS PM': 'I29', 'SEGURIDAD Y VIGILANCIA': 'I30', 'SERVICIOS INSTALACIONES': 'I31',
        'CELULARES': 'I32', 'SUMINISTROS GENERALES': 'I33', 'SUMINISTROS OFICINA': 'I34',
        'SUMINISTROS COMPUTO': 'I35', 'ARRENDAMIENTOS': 'I36', 'MANTENIMIENTOS': 'I37',
        'INVENTARIO FÍSICO': 'I38', 'OTROS IMPUESTOS Y DERECHOS': 'I39', 'NO DEDUCIBLES': 'I40',
        'SEGUROS Y FIANZAS': 'I41', 'CAPACITACION Y ENTRENAMIENTO': 'I42', 'MENSAJERIA': 'I43',
        'MUESTRAS': 'I44', 'FERIAS Y EXPOSICIONES': 'I45', 'PUBLICIDAD IMPRESA': 'I46',
        'IMPRESIONES 3D': 'I47', 'MATERIAL DISEÑO': 'I48', 'PATENTES': 'I49',
        'LICENCIAS Y SOFTWARE': 'I50', 'ATENCION A CLIENTES': 'I51', 'ASESORIAS PF': 'I52',
        'PORTALES CLIENTES': 'I53', 'CUOTAS Y SUSCRIPCIONES': 'I54', 'FLETES EXTERNOS': 'I55',
        'FLETES INTERNOS': 'I56', 'IMPTOS S/NOMINA': 'I57', 'CONTRIBUCIONES PATRONALES': 'I58',
        'TIMBRES Y FOLIOS FISCALES': 'I59', 'COMISION MERCANTIL': 'I60', 'GASTOS ADUANALES': 'I61'
    }

    gastos_operativos_subcuentas = {}
    for gasto, cell in gastos_operativos_map_cells.items():
        gastos_operativos_subcuentas[gasto] = {
            'actual': obtener_valor_celda(cell),
            'meta': META_VALUES['GASTOS_OPERATIVOS_INDIVIDUALES'].get(gasto, 0),
            'simulable': True
        }

    return {
        # Nivel 4 - VENTAS BRUTAS (INGRESO)
        'VENTAS BRUTAS': {
            'jerarquia': '4', 'tipo': 'suma',
            'componentes': ['VENTAS BRUTAS NACIONAL 16%', 'VENTAS BRUTAS EXTRANJERO'],
            'subcuentas': {
                'VENTAS BRUTAS NACIONAL 16%': {
                    'RETAIL': {'actual': obtener_valor_celda('I3'),
                               'meta': META_VALUES['VENTAS BRUTAS']['VENTAS BRUTAS NACIONAL 16%']['RETAIL'],
                               'simulable': True},
                    'CATALOGO': {'actual': obtener_valor_celda('I4'),
                                 'meta': META_VALUES['VENTAS BRUTAS']['VENTAS BRUTAS NACIONAL 16%']['CATALOGO'],
                                 'simulable': True},
                    'MAYOREO': {'actual': obtener_valor_celda('I5'),
                                'meta': META_VALUES['VENTAS BRUTAS']['VENTAS BRUTAS NACIONAL 16%']['MAYOREO'],
                                'simulable': True}
                },
                'VENTAS BRUTAS EXTRANJERO': {
                    'RETAIL': {'actual': obtener_valor_celda('I6'),
                               'meta': META_VALUES['VENTAS BRUTAS']['VENTAS BRUTAS EXTRANJERO']['RETAIL'],
                               'simulable': True},
                    'CATALOGO': {'actual': obtener_valor_celda('I7'),
                                 'meta': META_VALUES['VENTAS BRUTAS']['VENTAS BRUTAS EXTRANJERO']['CATALOGO'],
                                 'simulable': True},
                    'MAYOREO': {'actual': obtener_valor_celda('I8'),
                                'meta': META_VALUES['VENTAS BRUTAS']['VENTAS BRUTAS EXTRANJERO']['MAYOREO'],
                                'simulable': True}
                }
            }
        },

        # Nivel 5 - DESCUENTOS Y OTROS (DESCUENTOS son una reducción de ingresos, OTROS INGRESOS es un ingreso)
        'DESCUENTOS': {'jerarquia': '5', 'tipo': 'simple', 'actual': obtener_valor_celda('I9'),
                       'meta': META_VALUES['DESCUENTOS'],
                       'simulable': True},
        'OTROS INGRESOS': {'jerarquia': '5.3', 'tipo': 'simple', 'actual': obtener_valor_celda('I10'),
                             'meta': META_VALUES['OTROS INGRESOS'],
                             'simulable': True},

        # Nivel 6 - VENTAS NETAS (INGRESO CLAVE)
        'VENTAS NETAS': {'jerarquia': '6', 'tipo': 'formula', 'formula': 'VENTAS BRUTAS - DESCUENTOS + OTROS INGRESOS'},

        # Nivel 7 - COSTOS (EGRESO)
        'COSTO': {
            'jerarquia': '7', 'tipo': 'suma',
            'componentes': ['COSTO DIRECTO', 'COSTO INDIRECTO', 'OTROS COSTOS'],
            'subcuentas': {
                'COSTO DIRECTO': {
                    'MATERIALES A PROCESO': {'actual': obtener_valor_celda('I13'),
                                             'meta': META_VALUES['COSTO']['COSTO DIRECTO']['MATERIALES A PROCESO'],
                                             'simulable': True},
                    'MANO DE OBRA ARMADO': {'actual': obtener_valor_celda('I14'),
                                            'meta': META_VALUES['COSTO']['COSTO DIRECTO']['MANO DE OBRA ARMADO'],
                                            'simulable': True}
                },
                'COSTO INDIRECTO': {
                    'COSTOS DE CALIDAD': {'actual': obtener_valor_celda('I15'),
                                          'meta': META_VALUES['COSTO']['COSTO INDIRECTO']['COSTOS DE CALIDAD'],
                                          'simulable': True},
                    'COSTOS DE MOLDES': {'actual': obtener_valor_celda('I16'),
                                         'meta': META_VALUES['COSTO']['COSTO INDIRECTO']['COSTOS DE MOLDES'],
                                         'simulable': True}
                },
                'OTROS COSTOS': {
                    'OTROS COSTOS': {'actual': obtener_valor_celda('I17'), 'meta': META_VALUES['COSTO']['OTROS COSTOS'],
                                     'simulable': True}}
            }
        },

        # Nivel 8 - MARGEN BRUTO
        'MARGEN BRUTO': {'jerarquia': '8', 'tipo': 'formula', 'formula': 'VENTAS NETAS - COSTO'},

        # Nivel 9 - GASTOS OPERATIVOS (EGRESO)
        'TOTAL GASTOS OPERATIVOS': {
            'jerarquia': '9', 'tipo': 'suma_gastos',
            'subcuentas': gastos_operativos_subcuentas  # Se asigna el diccionario generado dinámicamente
        },

        # Resto de la estructura
        'EBITDA OPERATIVA': {'jerarquia': '10', 'tipo': 'formula',
                             'formula': 'VENTAS NETAS - COSTO - TOTAL GASTOS OPERATIVOS'},
        'TOTAL DE OTROS GASTOS': {'jerarquia': '11', 'tipo': 'simple', 'actual': obtener_valor_celda('I63'),
                                  'meta': META_VALUES['TOTAL DE OTROS GASTOS'],
                                  'simulable': True},
        'EBITDA': {'jerarquia': '12', 'tipo': 'formula', 'formula': 'EBITDA OPERATIVA - TOTAL DE OTROS GASTOS'},
        'FINANCIEROS': {
            'jerarquia': '13', 'tipo': 'suma',
            'componentes': ['GASTOS FINANCIEROS', 'PRODUCTOS FINANCIEROS', 'RESULTADO CAMBIARIO'],
            'subcuentas': {
                # Valores tomados directamente de la imagen para 'meta' y 'actual'
                'GASTOS FINANCIEROS': {'actual': obtener_valor_celda('I66'),
                                       'meta': META_VALUES['FINANCIEROS_INDIVIDUALES']['GASTOS FINANCIEROS'],
                                       'simulable': True},
                'PRODUCTOS FINANCIEROS': {'actual': obtener_valor_celda('I67'),
                                          'meta': META_VALUES['FINANCIEROS_INDIVIDUALES']['PRODUCTOS FINANCIEROS'],
                                          'simulable': True},
                'RESULTADO CAMBIARIO': {'actual': obtener_valor_celda('I68'),
                                        'meta': META_VALUES['FINANCIEROS_INDIVIDUALES']['RESULTADO CAMBIARIO'],
                                        'simulable': True}
            }
        },
        'BAI': {'jerarquia': '14', 'tipo': 'formula', 'formula': 'EBITDA - FINANCIEROS'}
    }

# --- FUNCIONES DE CÁLCULO ---
@st.cache_data
def get_cached_structure():
    return obtener_estructura_cuentas()

def get_actual_value(estructura, account_key, sub_account_key=None, sub_item_key=None):
    if sub_account_key and sub_item_key:
        return estructura.get(account_key, {}).get('subcuentas', {}).get(sub_account_key, {}).get(sub_item_key, {}).get(
            'actual', 0)
    elif sub_account_key:
        return estructura.get(account_key, {}).get('subcuentas', {}).get(sub_account_key, {}).get('actual', 0)
    else:
        return estructura.get(account_key, {}).get('actual', 0)

def get_meta_value(estructura, account_key, sub_account_key=None, sub_item_key=None):
    if sub_account_key and sub_item_key:
        return estructura.get(account_key, {}).get('subcuentas', {}).get(sub_account_key, {}).get(sub_item_key, {}).get(
            'meta', 0)
    elif sub_account_key:
        return estructura.get(account_key, {}).get('subcuentas', {}).get(sub_account_key, {}).get('meta', 0)
    else:
        return estructura.get(account_key, {}).get('meta', 0)

def calculate_account_value(_estructura, scenario, changes):
    results = {}

    def get_val(cuenta):
        if cuenta in results:
            return results[cuenta]

        datos = _estructura.get(cuenta)
        if not datos:
            return 0

        valor = 0
        if datos['tipo'] == 'simple':
            base_value = datos['actual'] if scenario == 'actual' else datos.get('meta', 0) if scenario == 'meta' else \
                datos['actual']
            if datos.get('simulable') and scenario == 'simulado':
                key = f"sim_{cuenta.replace(' ', '_').replace('/', '_')}"
                cambio_monetario = changes.get(key, 0.0)
                valor = base_value + cambio_monetario
            else:
                valor = base_value
        elif datos['tipo'] in ['suma', 'suma_gastos']:
            for subcuenta_name, subdatos_dict_or_item in datos['subcuentas'].items():
                if isinstance(subdatos_dict_or_item, dict) and 'actual' not in subdatos_dict_or_item:
                    for subitem_name, itemdatos in subdatos_dict_or_item.items():
                        base_value = itemdatos['actual'] if scenario == 'actual' else itemdatos.get('meta',
                                                                                                    0) if scenario == 'meta' else \
                            itemdatos['actual']
                        if itemdatos.get('simulable') and scenario == 'simulado':
                            key = f"sim_{subcuenta_name.replace(' ', '_')}_{subitem_name.replace(' ', '_')}"
                            cambio_monetario = changes.get(key, 0.0)
                            valor += base_value + cambio_monetario
                        else:
                            valor += base_value
                else:
                    itemdatos = subdatos_dict_or_item
                    base_value = itemdatos['actual'] if scenario == 'actual' else itemdatos.get('meta',
                                                                                                0) if scenario == 'meta' else \
                        itemdatos['actual']
                    if itemdatos.get('simulable') and scenario == 'simulado':
                        key = f"sim_{subcuenta_name.replace(' ', '_')}"
                        cambio_monetario = changes.get(key, 0.0)
                        valor += base_value + cambio_monetario
                    else:
                        valor += base_value
        elif datos['tipo'] == 'formula':
            temp_formula = datos['formula'].replace(' - ', '|').replace(' + ', '|')
            operands = temp_formula.split('|')
            operators = [char for char in datos['formula'] if char in ['+', '-']]

            first_operand_name = operands[0].strip()
            first_operand_val = results.get(first_operand_name)
            if first_operand_val is None:
                first_operand_val = get_val(first_operand_name)

            valor = first_operand_val
            for i, operand_name_raw in enumerate(operands[1:]):
                op = operators[i]
                current_operand_name = operand_name_raw.strip()
                current_operand_val = results.get(current_operand_name)
                if current_operand_val is None:
                    current_operand_val = get_val(current_operand_name)

                if op == '+':
                    valor += current_operand_val
                elif op == '-':
                    valor -= current_operand_val

        results[cuenta] = valor
        return valor

    cuentas_ordenadas = sorted(_estructura.keys(), key=lambda k: float(_estructura[k]['jerarquia']))

    for _ in range(5):
        for cuenta in cuentas_ordenadas:
            if cuenta not in results or _estructura[cuenta]['tipo'] == 'formula':
                get_val(cuenta)

    for cuenta in cuentas_ordenadas:
        if cuenta not in results:
            if _estructura[cuenta]['tipo'] == 'simple':
                results[cuenta] = _estructura[cuenta]['actual'] if scenario == 'actual' else _estructura[cuenta].get(
                    'meta', 0)
            elif _estructura[cuenta]['tipo'] in ['suma', 'suma_gastos']:
                sum_val = 0
                for subcuenta_name, subdatos_dict_or_item in _estructura[cuenta]['subcuentas'].items():
                    if isinstance(subdatos_dict_or_item, dict) and 'actual' not in subdatos_dict_or_item:
                        for subitem_name, itemdatos in subdatos_dict_or_item.items():
                            sum_val += itemdatos['actual'] if scenario == 'actual' else itemdatos.get('meta', 0)
                    else:
                        itemdatos = subdatos_dict_or_item
                        sum_val += itemdatos['actual'] if scenario == 'actual' else itemdatos.get('meta', 0)
                results[cuenta] = sum_val
    return results

def generar_dataframe_completo(changes):
    estructura = get_cached_structure()
    actual_values = calculate_account_value(estructura, 'actual', changes)
    meta_values = calculate_account_value(estructura, 'meta', changes)
    simulado_values = calculate_account_value(estructura, 'simulado', changes)

    data_list = []
    for c, i in estructura.items():
        data_list.append({
            'Cuenta': c,
            'Jerarquia': i['jerarquia'],
            'Actual': actual_values.get(c, 0),
            'Simulado': simulado_values.get(c, 0),
            'Meta': meta_values.get(c, 0)
        })
    df = pd.DataFrame(data_list)

    df['Jerarquia'] = pd.to_numeric(df['Jerarquia'])
    df = df.sort_values(by='Jerarquia').reset_index(drop=True)

    df['Brecha vs Meta (%)'] = ((df['Simulado'] - df['Meta']) / df['Meta'].replace(0, pd.NA)) * 100
    df['Brecha Simulado vs Actual (%)'] = ((df['Simulado'] - df['Actual']) / df['Actual'].replace(0, pd.NA)) * 100

    # --- LÓGICA PARA NUEVA COLUMNA ---
    ventas_netas_actual = df.loc[df['Cuenta'] == 'VENTAS NETAS', 'Actual'].iloc[0] if 'VENTAS NETAS' in df[
        'Cuenta'].values else 1
    ventas_netas_simulado = df.loc[df['Cuenta'] == 'VENTAS NETAS', 'Simulado'].iloc[0] if 'VENTAS NETAS' in df[
        'Cuenta'].values else 1
    ventas_netas_meta = df.loc[df['Cuenta'] == 'VENTAS NETAS', 'Meta'].iloc[0] if 'VENTAS NETAS' in df[
        'Cuenta'].values else 1
    
    df['Actual (% VN)'] = (df['Actual'] / ventas_netas_actual) * 100 if ventas_netas_actual != 0 else 0
    df['Simulado (% VN)'] = (df['Simulado'] / ventas_netas_simulado) * 100 if ventas_netas_simulado != 0 else 0
    df['Meta (% VN)'] = (df['Meta'] / ventas_netas_meta) * 100 if ventas_netas_meta != 0 else 0

    return df

def obtener_variables_modificadas(changes):
    variables_modificadas = []

    estructura = get_cached_structure()
    sim_key_info = {}
    for cuenta, datos in estructura.items():
        if datos.get('simulable') and 'subcuentas' not in datos:
            sim_key_info[f"sim_{cuenta.replace(' ', '_').replace('/', '_')}"] = {'name': cuenta,
                                                                                 'actual_val': datos['actual']}
        if 'subcuentas' in datos and isinstance(datos['subcuentas'], dict):
            for subcuenta_name, subdatos_value in datos['subcuentas'].items():
                if isinstance(subdatos_value, dict) and 'actual' in subdatos_value and subdatos_value.get('simulable'):
                    full_name = subcuenta_name
                    actual_val = subdatos_value['actual']
                    meta_val = subdatos_value['meta']
                    sim_key_info[f"sim_{subcuenta_name.replace(' ', '_')}"] = {
                        'name': full_name, 'actual_val': subdatos_value['actual']}
                elif isinstance(subdatos_value, dict) and 'actual' not in subdatos_value:
                    for subitem_name, itemdatos in subdatos_value.items():
                        if itemdatos.get('simulable'):
                            full_name = f"{subcuenta_name} - {subitem_name}"
                            actual_val = itemdatos['actual']
                            meta_val = itemdatos['meta']
                            sim_key_info[f"sim_{subcuenta_name.replace(' ', '_')}_{subitem_name.replace(' ', '_')}"] = {
                                'name': full_name, 'actual_val': itemdatos['actual']}

    for key, value in changes.items():
        if key.startswith('sim_') and value != 0.0 and key != 'sim_COSTO_DIRECTO_MATERIALES_A_PROCESO':
            info = sim_key_info.get(key)
            if info:
                display_name = info['name']
                actual_val = info['actual_val']

                porcentaje_cambio = 0.0
                if actual_val != 0:
                    porcentaje_cambio = (value / actual_val) * 100
                elif value != 0:
                    porcentaje_cambio = float('inf') if value > 0 else float('-inf')

                variables_modificadas.append({
                    'Variable': display_name,
                    'Cambio Monetario': value,
                    'Cambio Porcentual': porcentaje_cambio,
                    'ValorNumAbsoluto': abs(value)
                })
    if st.session_state.get('ajuste_activo', False):
        key_materiales_proceso = 'sim_COSTO_DIRECTO_MATERIALES_A_PROCESO'
        ajuste_val = changes.get(key_materiales_proceso, 0.0)
        actual_val_mp = get_actual_value(estructura, 'COSTO', 'COSTO DIRECTO', 'MATERIALES A PROCESO')
        if ajuste_val != 0:
            porcentaje_cambio_mp = 0.0
            if actual_val_mp != 0:
                porcentaje_cambio_mp = (ajuste_val / actual_val_mp) * 100
            elif ajuste_val != 0:
                porcentaje_cambio_mp = float('inf') if ajuste_val > 0 else float('-inf')

            variables_modificadas.append({
                'Variable': 'COSTO DIRECTO - MATERIALES A PROCESO (Ajuste Automático)',
                'Cambio Monetario': ajuste_val,
                'Cambio Porcentual': porcentaje_cambio_mp,
                'ValorNumAbsoluto': abs(ajuste_val)
            })

    df_variables = pd.DataFrame(variables_modificadas)
    if not df_variables.empty:
        df_variables = df_variables.sort_values('ValorNumAbsoluto', ascending=False)
        df_variables['Cambio Monetario'] = df_variables['Cambio Monetario'].apply(lambda x: f"${x:,.0f}")
        df_variables['Cambio Porcentual'] = df_variables['Cambio Porcentual'].apply(
            lambda x: f"{x:+.1f}%" if x != float('inf') and x != float('-inf') else (
                "+Inf%" if x == float('inf') else "-Inf%"))
        df_variables = df_variables.drop(columns=['ValorNumAbsoluto'])
    return df_variables

def aplicar_estilo_financiero(df):
    cuentas_ingresos = ['VENTAS BRUTAS', 'VENTAS NETAS', 'OTROS INGRESOS']
    cuentas_egresos = ['DESCUENTOS', 'COSTO', 'TOTAL GASTOS OPERATIVOS', 'TOTAL DE OTROS GASTOS', 'FINANCIEROS']
    cuentas_resultados = ['MARGEN BRUTO', 'EBITDA OPERATIVA', 'EBITDA', 'BAI']

    def estilo_fila(row):
        styles = [''] * len(row)
        if 'Cuenta' in row.index and row['Cuenta'] in (cuentas_ingresos + cuentas_egresos + cuentas_resultados):
            try:
                cuenta_idx = list(row.index).index('Cuenta')
                styles[cuenta_idx] = 'font-weight: bold;'
            except ValueError:
                pass
        if 'Brecha (% S vs M)' in row.index:
            brecha_val = row['Brecha (% S vs M)']
            try:
                brecha_idx = list(row.index).index('Brecha (% S vs M)')
            except ValueError:
                brecha_idx = -1

            if pd.notna(brecha_val) and brecha_idx != -1:
                if row['Cuenta'] in (cuentas_ingresos + cuentas_resultados):
                    if brecha_val > 0.01:
                        styles[brecha_idx] += f'color: {PALETA_GRAFICOS["Positivo"]}; font-weight: bold;'
                    elif brecha_val < -0.01:
                        styles[brecha_idx] += f'color: {PALETA_GRAFICOS["Negativo"]}; font-weight: bold;'
                elif row['Cuenta'] in cuentas_egresos:
                    if brecha_val < -0.01:
                        styles[brecha_idx] += f'color: {PALETA_GRAFICOS["Positivo"]}; font-weight: bold;'
                    elif brecha_val > 0.01:
                        styles[brecha_idx] += f'color: {PALETA_GRAFICOS["Negativo"]}; font-weight: bold;'
        return styles

    return df.style.apply(estilo_fila, axis=1)

def generar_recomendacion_variables_ia(df_completo):
    if not GEMINI_AVAILABLE:
        return "⚠️ **Error**: La API de Gemini no está configurada."

    estructura_original = get_cached_structure()
    simulable_accounts_details = []

    for cuenta_name, datos in estructura_original.items():
        if datos.get('simulable') and 'subcuentas' not in datos:
            actual_val = df_completo.loc[df_completo['Cuenta'] == cuenta_name, 'Actual'].iloc[0] if cuenta_name in \
                                                                                                   df_completo[
                                                                                                       'Cuenta'].values else \
                datos['actual']
            meta_val = df_completo.loc[df_completo['Cuenta'] == cuenta_name, 'Meta'].iloc[0] if cuenta_name in \
                                                                                               df_completo[
                                                                                                   'Cuenta'].values else \
                datos['meta']
            simulable_accounts_details.append({
                'Variable': cuenta_name, 'Actual': actual_val, 'Meta': meta_val
            })
        elif 'subcuentas' in datos and isinstance(datos['subcuentas'], dict):
            for subcuenta_name, subdatos in datos['subcuentas'].items():
                if isinstance(subdatos, dict) and 'actual' in subdatos and subdatos.get('simulable'):
                    full_name = subcuenta_name
                    actual_val = subdatos['actual']
                    meta_val = subdatos['meta']
                    simulable_accounts_details.append({
                        'Variable': full_name, 'Actual': actual_val, 'Meta': meta_val
                    })
                elif isinstance(subdatos, dict) and 'actual' not in subdatos:
                    for subitem_name, itemdatos in subdatos.items():
                        if itemdatos.get('simulable'):
                            full_name = f"{subcuenta_name} - {subitem_name}"
                            actual_val = itemdatos['actual']
                            meta_val = itemdatos['meta']
                            simulable_accounts_details.append({
                                'Variable': full_name, 'Actual': actual_val, 'Meta': meta_val
                            })

    df_simulable = pd.DataFrame(simulable_accounts_details)
    df_simulable['Desviacion (Actual vs Meta)'] = df_simulable['Actual'] - df_simulable['Meta']
    df_simulable['Abs_Deviation'] = df_simulable['Desviacion (Actual vs Meta)'].abs()

    df_top_deviations = df_simulable[df_simulable['Abs_Deviation'] > 1000].sort_values(by='Abs_Deviation',
                                                                                       ascending=False).head(7)

    top_deviations_table_md = "No se identificaron desviaciones significativas entre el Actual y la Meta para recomendar acciones en este momento."
    if not df_top_deviations.empty:
        top_deviations_table_md = "A continuación, se presenta una tabla con las **variables que muestran las mayores desviaciones monetarias entre su valor Actual y la Meta**. Estas son las áreas clave recomendadas para enfocar tus esfuerzos de simulación, ya que representan el mayor potencial de mejora o riesgo:\n\n"
        top_deviations_table_md += "| Variable | Actual | Meta | Desviación (Actual vs Meta) |\n"
        top_deviations_table_md += "|:---------|-------:|-----:|----------------------------:|\n"
        for _, row in df_top_deviations.iterrows():
            top_deviations_table_md += f"| {row['Variable']} | ${row['Actual']:,.0f} | ${row['Meta']:,.0f} | ${row['Desviacion (Actual vs Meta)']:+,.0f} |\n"
        top_deviations_table_md += "\n"

    company_context = """
    La empresa es **PORTAWARE**, fabricante de artículos para el hogar, predominantemente de plástico. Tienen fuertes expectativas de crecimiento a nivel nacional y están comenzando a expandirse en mercados internacionales. El ambiente económico actual es volátil, con presiones inflacionarias en materias primas (plástico, derivados del petróleo) y fluctuaciones en las tasas de cambio. La estrategia de la empresa debe enfocarse en la eficiencia operativa, la gestión de costos, y la optimización de ingresos en un entorno de expansión.
    """

    prompt = f"""
    Eres un Director Financiero (CFO) experto, muy conciso y enfocado en la estrategia. Tu tarea es analizar las desviaciones entre el desempeño financiero **Actual** y la **Meta** establecida para PORTAWARE. Luego, identificar las variables más críticas para que un usuario las mueva en un simulador financiero, siendo muy concreto y ejecutivo en tu recomendación.

    **Contexto de PORTAWARE:**
    {company_context}

    {top_deviations_table_md}

    **Análisis y Recomendación Estratégica (máximo 100 palabras):**
    Basándote únicamente en la tabla de desviaciones anterior y el contexto de PORTAWARE:
    1.  **Diagnóstico Rápido:** ¿Cuál es la tendencia general de estas desviaciones y su implicación para el EBITDA? ¿Estamos por encima o por debajo de la meta en las áreas clave?
    2.  **Variables Clave y Dirección de Ajuste:** Para las 3-5 variables con mayor desviación (no más de 5):
        * Nombra la variable.
        * Indica brevemente el impacto en el BAI y la conexión con el contexto de PORTAWARE (ej., "impacto directo en ventas y la estrategia de expansión", "reduce margen por costos de materia prima").
        * Menciona una **acción estratégica específica y cuantificable** (si es posible) para esa variable, considerando el mercado y ambiente económico actual.
    3.  **Priorización:** ¿Cuáles 2-3 variables (revisa sub cuenta) de esta lista deberían ser la *máxima prioridad* para la simulación inicial, y por qué, en línea con el crecimiento nacional e internacional de PORTAWARE?
    4.  **Conclusión General:** Una breve frase de cierre sobre el potencial de mejora estratégica.

    Usa un tono profesional y directo. Formatea tu respuesta con Markdown, incluyendo negritas y listas.
    """
    try:
        model = genai.GenerativeModel('gemini-1.5-flash')
        response = model.generate_content(prompt)
        return response.text
    except Exception as e:
        return f"❌ **Error al contactar la API de Gemini**: {e}"

def generar_insight_financiero(df_completo, actual_col='Actual', meta_col='Meta', simulado_col='Simulado'):
    if not GEMINI_AVAILABLE:
        return "⚠️ **Error**: La API de Gemini no está configurada."

    df_analisis = df_completo.copy()
    df_analisis['Brecha Actual vs Meta'] = df_analisis[actual_col] - df_analisis[meta_col]
    df_analisis['Brecha Simulado vs Actual'] = df_analisis[simulado_col] - df_analisis[actual_col]
    df_analisis['Brecha Simulado vs Meta'] = df_analisis[simulado_col] - df_analisis[meta_col]

    cuentas_para_ia = df_analisis.sort_values(by='Jerarquia').reset_index(drop=True)

    razones_data = []

    ventas_netas_actual = cuentas_para_ia.loc[cuentas_para_ia['Cuenta'] == 'VENTAS NETAS', actual_col].iloc[
        0] if 'VENTAS NETAS' in cuentas_para_ia['Cuenta'].values else 1
    margen_bruto_actual = cuentas_para_ia.loc[cuentas_para_ia['Cuenta'] == 'MARGEN BRUTO', actual_col].iloc[
        0] if 'MARGEN BRUTO' in cuentas_para_ia['Cuenta'].values else 0
    ebitda_actual = cuentas_para_ia.loc[cuentas_para_ia['Cuenta'] == 'EBITDA', actual_col].iloc[0] if 'EBITDA' in \
                                                                                                     cuentas_para_ia[
                                                                                                         'Cuenta'].values else 0
    bai_actual = cuentas_para_ia.loc[cuentas_para_ia['Cuenta'] == 'BAI', actual_col].iloc[0] if 'BAI' in \
                                                                                               cuentas_para_ia[
                                                                                                   'Cuenta'].values else 0

    ventas_netas_meta = cuentas_para_ia.loc[cuentas_para_ia['Cuenta'] == 'VENTAS NETAS', meta_col].iloc[
        0] if 'VENTAS NETAS' in cuentas_para_ia['Cuenta'].values else 1
    margen_bruto_meta = cuentas_para_ia.loc[cuentas_para_ia['Cuenta'] == 'MARGEN BRUTO', meta_col].iloc[
        0] if 'MARGEN BRUTO' in cuentas_para_ia['Cuenta'].values else 0
    ebitda_meta = cuentas_para_ia.loc[cuentas_para_ia['Cuenta'] == 'EBITDA', meta_col].iloc[0] if 'EBITDA' in \
                                                                                                 cuentas_para_ia[
                                                                                                     'Cuenta'].values else 0
    bai_meta = cuentas_para_ia.loc[cuentas_para_ia['Cuenta'] == 'BAI', meta_col].iloc[0] if 'BAI' in cuentas_para_ia[
        'Cuenta'].values else 1

    ventas_netas_simulado = cuentas_para_ia.loc[cuentas_para_ia['Cuenta'] == 'VENTAS NETAS', simulado_col].iloc[
        0] if 'VENTAS NETAS' in cuentas_para_ia['Cuenta'].values else 1
    margen_bruto_simulado = cuentas_para_ia.loc[cuentas_para_ia['Cuenta'] == 'MARGEN BRUTO', simulado_col].iloc[
        0] if 'MARGEN BRUTO' in cuentas_para_ia['Cuenta'].values else 0
    ebitda_simulado = cuentas_para_ia.loc[cuentas_para_ia['Cuenta'] == 'EBITDA', simulado_col].iloc[0] if 'EBITDA' in \
                                                                                                         cuentas_para_ia[
                                                                                                             'Cuenta'].values else 0
    bai_simulado = cuentas_para_ia.loc[cuentas_para_ia['Cuenta'] == 'BAI', simulado_col].iloc[0] if 'BAI' in \
                                                                                                   cuentas_para_ia[
                                                                                                       'Cuenta'].values else 0

    margen_bruto_vn_actual = (margen_bruto_actual / ventas_netas_actual * 100) if ventas_netas_actual != 0 else 0
    margen_bruto_vn_meta = (margen_bruto_meta / ventas_netas_meta * 100) if ventas_netas_meta != 0 else 0
    margen_bruto_vn_simulado = (
                                   margen_bruto_simulado / ventas_netas_simulado * 100) if ventas_netas_simulado != 0 else 0
    razones_data.append({'Razon Financiera': 'Margen Bruto sobre Ventas Netas (%)', 'Actual': margen_bruto_vn_actual,
                         'Meta': margen_bruto_vn_meta, 'Simulado': margen_bruto_vn_simulado})

    margen_ebitda_vn_actual = (ebitda_actual / ventas_netas_actual * 100) if ventas_netas_actual != 0 else 0
    margen_ebitda_vn_meta = (ebitda_meta / ventas_netas_meta * 100) if ventas_netas_meta != 0 else 0
    margen_ebitda_vn_simulado = (ebitda_simulado / ventas_netas_simulado * 100) if ventas_netas_simulado != 0 else 0
    razones_data.append({'Razon Financiera': 'Margen EBITDA sobre Ventas Netas (%)', 'Actual': margen_ebitda_vn_actual,
                         'Meta': margen_ebitda_vn_meta, 'Simulado': margen_ebitda_vn_simulado})

    margen_bai_vn_actual = (bai_actual / ventas_netas_actual * 100) if ventas_netas_actual != 0 else 0
    margen_bai_vn_meta = (bai_meta / ventas_netas_meta * 100) if ventas_netas_meta != 0 else 0
    margen_bai_vn_simulado = (bai_simulado / ventas_netas_simulado * 100) if ventas_netas_simulado != 0 else 0
    razones_data.append({'Razon Financiera': 'Margen BAI sobre Ventas Netas (%)', 'Actual': margen_bai_vn_actual,
                         'Meta': margen_bai_vn_meta, 'Simulado': margen_bai_vn_simulado})

    df_razones = pd.DataFrame(razones_data)
    df_razones['Brecha Actual vs Meta (%)'] = df_razones['Actual'] - df_razones['Meta']
    df_razones['Brecha Simulado vs Actual (%)'] = df_razones['Simulado'] - df_razones['Actual']
    df_razones['Brecha Simulado vs Meta (%)'] = df_razones['Simulado'] - df_razones['Meta']

    cols_for_prompt_df = cuentas_para_ia[['Cuenta', actual_col, meta_col, simulado_col,
                                          'Brecha Actual vs Meta', 'Brecha Simulado vs Actual',
                                          'Brecha Simulado vs Meta']].copy()

    for col in [actual_col, meta_col, simulado_col, 'Brecha Actual vs Meta', 'Brecha Simulado vs Actual',
                'Brecha Simulado vs Meta']:
        if col in cols_for_prompt_df.columns:
            cols_for_prompt_df[col] = pd.to_numeric(cols_for_prompt_df[col], errors='coerce').fillna(0)
            cols_for_prompt_df[col] = cols_for_prompt_df[col].apply(lambda x: f"{x:+.0f}")

    razones_table_md = df_razones.to_markdown(index=False, floatfmt="+.2f")
    analysis_table_md = cols_for_prompt_df.to_markdown(index=False, numalign="left", stralign="left")

    company_context = """
    La empresa es **PORTAWARE**, fabricante de artículos para el hogar, predominantemente de plástico. Tienen fuertes expectativas de crecimiento a nivel nacional y están comenzando a expandirse en mercados internacionales. El ambiente económico actual es volátil, con presiones inflacionarias en materias primas (plástico, derivados del petróleo) y fluctuaciones en las tasas de cambio. La estrategia de la empresa debe enfocarse en la eficiencia operativa, la gestión de costos, y la optimización de ingresos en un entorno de expansión.
    """

    top_deviations_context_for_insight = st.session_state.get('recomendacion_ia_dashboard_content',
                                                              "*(No se ha generado una tabla de desviaciones de variables clave aún. Haz clic en 'Obtener Recomendación IA de Variables Clave' en el Dashboard de Brechas para verla.)*"
                                                              )

    prompt = f"""
    Eres un Director Financiero (CFO) experto, muy conciso y estratégico. Tu tarea es analizar el desempeño financiero de la empresa PORTAWARE comparando el escenario **Actual** con la **Meta** establecida, considerando también el escenario **Simulado** (si hay cambios). Proporciona un diagnóstico ejecutivo y recomendaciones estratégicas concretas, integrando el contexto de la empresa, el mercado, el ambiente económico y las razones financieras clave.

    **Contexto de PORTAWARE:**
    {company_context}

    **Variables Clave con Mayores Desviaciones (si disponibles):**
    {top_deviations_context_for_insight}

    **Datos Financieros Clave (Valores Absolutos):**
    {analysis_table_md}

    **Razones Financieras (Rentabilidad y Eficiencia):**
    {razones_table_md}

    **Análisis Estratégico y Recomendaciones Ejecutivas (Máximo 100 palabras):**

    1.  **Panorama General y Razones Clave (Actual vs. Meta y Simulado):**
        -   Inicia con un resumen de 1-2 frases sobre el cumplimiento de la meta del BAI y las principales tendencias en las **razones financieras **. ¿Cómo se comparan los porcentajes Actuales, Meta y Simulados? ¿Qué implicaciones tiene para la salud financiera y la estrategia de PORTAWARE?
        -   Identifica las 2-3 áreas principales (Ventas, Costos, Gastos) y las razones financieras asociadas que explican la mayor parte de la desviación del BAI y los márgenes.

    2.  **Recomendaciones Estratégicas y Cuantificables:**
        -   Proporciona 2-3 recomendaciones *accionables* para PORTAWARE, enfocadas en mejorar las razones financieras y el BAI, priorizando el cierre de las brechas más grandes o el impulso de la estrategia de expansión. Incluye:
            * **Acción específica** (ej: "Optimizar la compra de polímeros para mejorar el Margen Bruto en X puntos porcentuales").
            * **Razón Financiera impactada** y su potencial de mejora.
            * **Relevancia estratégica** para PORTAWARE, considerando su expansión internacional y el manejo de costos.

    3.  **Conclusión Ejecutiva:**
        -   Una breve declaración sobre la visión general, el potencial de mejora, y los próximos pasos estratégicos para PORTAWARE, con énfasis in the profitability and operational efficiency on its path to growth.

    Sé conciso, directo y enfocado en la toma de decisiones estratégicas de alto nivel. Usa negritas para resaltar conceptos clave, cursivas para énfasis y listas para las recomendaciones.
    """
    try:
        model = genai.GenerativeModel('gemini-1.5-flash')
        response = model.generate_content(prompt)
        return response.text
    except Exception as e:
        return f"❌ **Error al contactar la API de Gemini**: {e}"


# --- FUNCIONES DE CALLBACK PARA LA INTERFAZ ---

def save_current_scenario_callback():
    scenario_name = st.session_state.new_scenario_name_input
    if scenario_name:
        # Recopila solo los cambios realizados por el usuario (diferentes de cero)
        current_changes = {
            key: value for key, value in st.session_state.items()
            if key.startswith('sim_') and value != 0.0
        }
        # Filtra el ajuste automático para no guardarlo como una decisión manual
        current_changes_filtered = {
            k: v for k, v in current_changes.items()
            if not k.startswith('sim_COSTO_DIRECTO_MATERIALES_A_PROCESO')
        }

        st.session_state['saved_scenarios'][scenario_name] = current_changes_filtered
        save_scenarios_to_file()
        st.success(f"Escenario '{scenario_name}' guardado.")
        st.session_state.new_scenario_name_input = "" # Limpia el campo de texto
    else:
        st.warning("Por favor, introduce un nombre para el escenario.")

def load_scenario_callback():
    scenario_name = st.session_state.load_scenario_selectbox
    if scenario_name and scenario_name in st.session_state['saved_scenarios']:
        loaded_changes = st.session_state['saved_scenarios'][scenario_name]
        
        # Primero, resetea todos los valores de simulación a cero
        for key in list(st.session_state.keys()):
            if key.startswith('sim_'):
                st.session_state[key] = 0.0
        
        # Luego, aplica los valores del escenario cargado
        for key, value in loaded_changes.items():
            st.session_state[key] = value

        # Resetea los controles de ajuste automático al cargar un escenario
        st.session_state['ajuste_activo'] = False
        st.session_state['porcentaje_ajuste'] = 45

        st.success(f"Escenario '{scenario_name}' cargado.")
    elif scenario_name:
        st.error(f"No se pudo cargar el escenario '{scenario_name}'.")

def delete_scenario_callback():
    scenario_name = st.session_state.delete_scenario_selectbox
    if scenario_name and scenario_name in st.session_state.get('saved_scenarios', {}):
        del st.session_state['saved_scenarios'][scenario_name]
        save_scenarios_to_file()
        st.success(f"Escenario '{scenario_name}' eliminado.")
        # Limpia la selección para evitar re-eliminación accidental
        st.session_state.delete_scenario_selectbox = ""
    elif scenario_name:
        st.error(f"No se pudo eliminar el escenario '{scenario_name}'.")

def reset_simulator_callback():
    # Resetea todos los valores de simulación a cero
    for key in list(st.session_state.keys()):
        if key.startswith('sim_'):
            st.session_state[key] = 0.0
    
    # Resetea los controles y los informes de la IA
    st.session_state['ajuste_activo'] = False
    st.session_state['porcentaje_ajuste'] = 45
    if 'recomendacion_ia_dashboard' in st.session_state:
        del st.session_state['recomendacion_ia_dashboard']
    if 'recomendacion_ia_dashboard_content' in st.session_state:
        del st.session_state['recomendacion_ia_dashboard_content']
    if 'informe_ia' in st.session_state:
        del st.session_state['informe_ia']
    st.success("Simulador reseteado a los valores iniciales.")

def simulate_plus_10_percent_sales_callback():
    estructura = get_cached_structure()
    
    # Define los canales de ventas nacionales
    canales_nacionales = ['RETAIL', 'CATALOGO', 'MAYOREO']
    
    for canal in canales_nacionales:
        actual_val = get_actual_value(estructura, 'VENTAS BRUTAS', 'VENTAS BRUTAS NACIONAL 16%', canal)
        key = f"sim_VENTAS_BRUTAS_NACIONAL_16%_{canal}"
        # El cambio es el 10% del valor actual. Se SUMA al cambio existente.
        st.session_state[key] += actual_val * 0.10
    st.success("Se aplicó un +10% a las ventas nacionales.")

# --- FLUJO PRINCIPAL DE LA APLICACIÓN ---

# Título de la aplicación
st.title('📊 Simulador Financiero Jerárquico')

# Inicializar estado de la sesión y cargar datos de Excel si es necesario
if 'excel_data' not in st.session_state:
    if not cargar_excel():
        st.error("Error crítico al cargar Excel. La aplicación no puede continuar.")
        st.stop()

initialize_session_state()

st.caption(
    f"Estado de IA: {'✅ Conectada' if GEMINI_AVAILABLE else '❌ No disponible'}"
)

# --- CÁLCULOS PRINCIPALES ---

# Obtener cambios actuales del session_state
changes = {key: value for key, value in st.session_state.items() if key.startswith('sim_')}

# Lógica para ajuste automático de costos
key_materiales_proceso = 'sim_COSTO_DIRECTO_MATERIALES_A_PROCESO'
if st.session_state.get('ajuste_activo', False):
    cambio_ventas_brutas_nacional = 0
    for canal in ['RETAIL', 'CATALOGO', 'MAYOREO']:
        key_ventas_nacional = f"sim_VENTAS_BRUTAS_NACIONAL_16%_{canal}"
        cambio_ventas_brutas_nacional += max(0, st.session_state.get(key_ventas_nacional, 0.0))

    porcentaje_ajuste = st.session_state.get('porcentaje_ajuste', 45)
    ajuste_materiales = (porcentaje_ajuste / 100) * cambio_ventas_brutas_nacional
    # Este cambio se aplica directamente al diccionario 'changes' para el cálculo
    changes[key_materiales_proceso] = ajuste_materiales
else:
    # Si el ajuste no está activo, el cambio manual (si existe) se usa, o cero.
    changes[key_materiales_proceso] = st.session_state.get(key_materiales_proceso, 0.0)


# Generar DataFrames con los cambios actualizados
df_completo = generar_dataframe_completo(changes)
df_variables_mod = obtener_variables_modificadas(changes)


# --- BARRA LATERAL (SIDEBAR) ---
with st.sidebar:
    st.header("⚙️ Controles de Simulación")
    st.info(
        "**IMPORTANTE:** Los escenarios guardados son temporales en la versión web. "
        "Para persistencia, se necesita una base de datos."
    )
    
    # Sección de Gestión de Escenarios
    st.subheader("💾 Gestión de Escenarios")
    scenario_names = list(st.session_state.get('saved_scenarios', {}).keys())

    col1, col2 = st.columns([3, 1])
    with col1:
        st.text_input("Nombre del nuevo escenario:", key="new_scenario_name_input", placeholder="Ej: Proyección optimista")
    with col2:
        st.button("Guardar", use_container_width=True, on_click=save_current_scenario_callback, key="save_btn")

    if scenario_names:
        st.selectbox("Cargar escenario:", [""] + scenario_names, key="load_scenario_selectbox", on_change=load_scenario_callback)
        st.selectbox("Eliminar escenario:", [""] + scenario_names, key="delete_scenario_selectbox", on_change=delete_scenario_callback)
    else:
        st.caption("No hay escenarios guardados.")

    st.markdown("---")
    
    # Escenarios rápidos
    st.subheader("⚡ Acciones Rápidas")
    col1, col2 = st.columns(2)
    with col1:
        st.button("Resetear Simulador", use_container_width=True, on_click=reset_simulator_callback)
    with col2:
        st.button("+10% Ventas Nac.", use_container_width=True, on_click=simulate_plus_10_percent_sales_callback)

    if st.button("Recargar Datos de Excel", use_container_width=True):
        with st.spinner("Recargando..."):
            st.cache_data.clear()
            cargar_excel()
            st.success("Datos de Excel recargados.")
            time.sleep(1)
            st.rerun()
    if 'last_modified' in st.session_state:
        st.caption(f"Última recarga: {st.session_state.last_modified.strftime('%H:%M:%S')}")
    
    st.markdown("---")

    # Ajuste automático
    st.subheader("📦 Ajuste Automático de Costos")
    st.checkbox("Activar ajuste de costo de material", key='ajuste_activo')
    if st.session_state.ajuste_activo:
        st.slider(
            "Ajuste sobre ↑ Ventas Nac. (%)", 0, 100,
            key='porcentaje_ajuste',
            help="El costo de 'Materiales a Proceso' aumentará en este porcentaje del AUMENTO en Ventas Brutas Nacionales."
        )

    st.markdown("---")
    
    # Pestañas de controles de simulación
    estructura = get_cached_structure()
    tab_ventas, tab_costos, tab_gastos, tab_otros = st.tabs(["💰", "🏭", "💸", "📊"])

    # ... (El código de la UI dentro de las pestañas de la sidebar va aquí)
    # ... (No se necesitan cambios en esta sección, se omite por brevedad)
    def display_number_input_info_with_actual_meta_brecha(key, actual_val, meta_val, is_auto_adjusted=False):
        current_change = st.session_state.get(key, 0.0)
        percentage_change_sim = 0.0
        if actual_val != 0:
            percentage_change_sim = (current_change / actual_val) * 100
        elif current_change != 0:
            percentage_change_sim = float('inf') if current_change > 0 else float('-inf')

        simulated_val_for_brecha = actual_val + current_change
        brecha_vs_meta = simulated_val_for_brecha - meta_val
        brecha_vs_meta_percent = 0.0
        if meta_val != 0:
            brecha_vs_meta_percent = (brecha_vs_meta / meta_val) * 100
        elif brecha_vs_meta != 0:
            brecha_vs_meta_percent = float('inf') if brecha_vs_meta > 0 else float('-inf')

        st.markdown(
            f'<div class="small-input-info">'
            f'<p><b>Actual:</b> ${actual_val:,.0f} | '
            f'<b>Meta:</b> ${meta_val:,.0f} | '
            f'<b>Brecha vs Meta:</b> {brecha_vs_meta_percent:+.1f}%</p>'
            f'<p><b>{"Ajuste Automático" if is_auto_adjusted else "Cambio"}:</b> ${current_change:,.0f} ({percentage_change_sim:+.1f}%)</p>'
            f'</div>',
            unsafe_allow_html=True
        )

    with tab_ventas:
        st.subheader("Ventas por Canal")
        with st.expander("🏠 Ventas Nacionales", expanded=True):
            for canal, datos in estructura['VENTAS BRUTAS']['subcuentas']['VENTAS BRUTAS NACIONAL 16%'].items():
                key = f"sim_VENTAS_BRUTAS_NACIONAL_16%_{canal}"
                actual_val = get_actual_value(estructura, 'VENTAS BRUTAS', 'VENTAS BRUTAS NACIONAL 16%', canal)
                meta_val = get_meta_value(estructura, 'VENTAS BRUTAS', 'VENTAS BRUTAS NACIONAL 16%', canal)
                min_val = -float(actual_val * 2) if actual_val > 0 else -200000000.0
                max_val = float(actual_val * 2) if actual_val > 0 else 200000000.0
                st.number_input(f"{canal}", min_value=min_val, max_value=max_val, value=st.session_state.get(key, 0.0),
                                step=10000.0, key=key)
                display_number_input_info_with_actual_meta_brecha(key, actual_val, meta_val)

        with st.expander("🌎 Ventas Extranjero"):
            for canal, datos in estructura['VENTAS BRUTAS']['subcuentas']['VENTAS BRUTAS EXTRANJERO'].items():
                key = f"sim_VENTAS_BRUTAS_EXTRANJERO_{canal}"
                actual_val = get_actual_value(estructura, 'VENTAS BRUTAS', 'VENTAS BRUTAS EXTRANJERO', canal)
                meta_val = get_meta_value(estructura, 'VENTAS BRUTAS', 'VENTAS BRUTAS EXTRANJERO', canal)
                min_val = -float(actual_val * 2) if actual_val > 0 else -2000000.0
                max_val = float(actual_val * 2) if actual_val > 0 else 2000000.0
                st.number_input(f"{canal} (Ext)", min_value=min_val, max_value=max_val,
                                value=st.session_state.get(key, 0.0), step=1000.0, key=key)
                display_number_input_info_with_actual_meta_brecha(key, actual_val, meta_val)

        key = 'sim_DESCUENTOS'
        actual_desc = get_actual_value(estructura, 'DESCUENTOS')
        meta_desc = get_meta_value(estructura, 'DESCUENTOS')
        min_desc = -float(actual_desc * 2) if actual_desc > 0 else -10000000.0
        max_desc = float(actual_desc * 2) if actual_desc > 0 else 10000000.0
        st.number_input("Descuentos", min_value=min_desc, max_value=max_desc, value=st.session_state.get(key, 0.0),
                        step=5000.0, key=key)
        display_number_input_info_with_actual_meta_brecha(key, actual_desc, meta_desc)

        key = 'sim_OTROS_INGRESOS'
        actual_otros_ingresos = get_actual_value(estructura, 'OTROS INGRESOS')
        meta_otros_ingresos = get_meta_value(estructura, 'OTROS INGRESOS')
        min_oi = -float(abs(actual_otros_ingresos) * 5) if actual_otros_ingresos != 0 else -5000000.0
        max_oi = float(abs(actual_otros_ingresos) * 5) if actual_otros_ingresos != 0 else 5000000.0
        st.number_input("Otros Ingresos", min_value=min_oi, max_value=max_oi, value=st.session_state.get(key, 0.0),
                        step=100.0, key=key)
        display_number_input_info_with_actual_meta_brecha(key, actual_otros_ingresos, meta_otros_ingresos)

    with tab_costos:
        st.subheader("Estructura de Costos")
        with st.expander("🎯 Costos Directos", expanded=True):
            for item, datos in estructura['COSTO']['subcuentas']['COSTO DIRECTO'].items():
                key = f"sim_COSTO_DIRECTO_{item.replace(' ', '_')}"
                actual_val = get_actual_value(estructura, 'COSTO', 'COSTO DIRECTO', item)
                meta_val = get_meta_value(estructura, 'COSTO', 'COSTO DIRECTO', item)
                min_val = -float(actual_val * 2) if actual_val > 0 else -50000000.0
                max_val = float(actual_val * 2) if actual_val > 0 else 50000000.0

                if item == 'MATERIALES A PROCESO':
                    if st.session_state.get('ajuste_activo', False):
                        # El valor mostrado se basa en el cálculo, pero el input está deshabilitado
                        # y el valor real se maneja en el diccionario 'changes'
                        st.number_input(
                            label=f"{item.replace('_', ' ').title()}",
                            value=changes.get(key, 0.0), # Muestra el ajuste calculado
                            key=f"display_{key}", # Usa una clave diferente para no interferir
                            disabled=True,
                            help="Ajuste automático activado. Se calcula basado en ventas."
                        )
                        # Aquí la clave real 'key' se usa para obtener el valor del estado
                        display_number_input_info_with_actual_meta_brecha(key, actual_val, meta_val, is_auto_adjusted=True)
                    else:
                        # El usuario puede editarlo manualmente si el ajuste está apagado
                        st.number_input(f"{item.replace('_', ' ').title()}", min_value=min_val, max_value=max_val,
                                        value=st.session_state.get(key, 0.0), step=1000.0, key=key)
                        display_number_input_info_with_actual_meta_brecha(key, actual_val, meta_val)
                else:
                    st.number_input(f"{item.replace('_', ' ').title()}", min_value=min_val, max_value=max_val,
                                    value=st.session_state.get(key, 0.0), step=1000.0, key=key)
                    display_number_input_info_with_actual_meta_brecha(key, actual_val, meta_val)

    with tab_gastos:
        #... tu código para la pestaña de gastos ...
        pass

    with tab_otros:
        #... tu código para la pestaña de otros ...
        pass
        
# --- CONTENIDO PRINCIPAL ---
tab1, tab2 = st.tabs(["📊 Dashboard de Brechas", "📈 Análisis Visual e IA"])

with tab1:
    # ... (Tu código para el dashboard principal va aquí)
    # ... (No se necesitan cambios en esta sección, se omite por brevedad)
    st.header("Dashboard de Brechas vs. Meta")
    st.info("💡 **Consejo:** El botón (☰) en la esquina superior izquierda ahora te permite mostrar u ocultar los controles de simulación.")

    col1, col2, col3, col4 = st.columns(4)
    ventas_netas = df_completo.loc[df_completo['Cuenta'] == 'VENTAS NETAS', 'Simulado'].iloc[0]
    margen_bruto = df_completo.loc[df_completo['Cuenta'] == 'MARGEN BRUTO', 'Simulado'].iloc[0]
    ebitda = df_completo.loc[df_completo['Cuenta'] == 'EBITDA', 'Simulado'].iloc[0]
    bai = df_completo.loc[df_completo['Cuenta'] == 'BAI', 'Simulado'].iloc[0]

    ventas_netas_meta = df_completo.loc[df_completo['Cuenta'] == 'VENTAS NETAS', 'Meta'].iloc[0]
    margen_bruto_meta = df_completo.loc[df_completo['Cuenta'] == 'MARGEN BRUTO', 'Meta'].iloc[0]
    ebitda_meta = df_completo.loc[df_completo['Cuenta'] == 'EBITDA', 'Meta'].iloc[0]
    bai_meta = df_completo.loc[df_completo['Cuenta'] == 'BAI', 'Meta'].iloc[0]

    diff_ventas_netas = ventas_netas - ventas_netas_meta
    diff_margen_bruto = margen_bruto - margen_bruto_meta
    diff_ebitda = ebitda - ebitda_meta
    diff_bai = bai - bai_meta

    col1.metric("VENTAS NETAS", f"${ventas_netas:,.0f}", f"{diff_ventas_netas:+,.0f} vs Meta",
                delta_color="normal")
    col2.metric("MARGEN BRUTO", f"${margen_bruto:,.0f}", f"{diff_margen_bruto:+,.0f} vs Meta",
                delta_color="normal")
    col3.metric("EBITDA", f"${ebitda:,.0f}", f"{diff_ebitda:+,.0f} vs Meta",
                delta_color="normal")
    col4.metric("BAI", f"${bai:,.0f}", f"{diff_bai:+,.0f} vs Meta",
                delta_color="normal")

    st.subheader("Análisis Comparativo Detallado")
    
    # --- LÓGICA DE VISUALIZACIÓN DE LA TABLA CON LA NUEVA COLUMNA ---
    df_display = df_completo[[
        'Cuenta', 'Actual', 'Actual (% VN)', 'Simulado', 'Simulado (% VN)', 'Meta', 'Meta (% VN)', 'Brecha vs Meta (%)'
    ]].copy()
    
    df_display.rename(columns={
        'Actual (% VN)': '% Act. vs VN',
        'Simulado (% VN)': '% Sim. vs VN',
        'Meta (% VN)': '% Meta vs VN',
        'Brecha vs Meta (%)': 'Brecha (% S vs M)'
    }, inplace=True)

    formatos = {
        'Actual': '{:,.0f}',
        '% Act. vs VN': '{:,.2f}%',
        'Simulado': '{:,.0f}',
        '% Sim. vs VN': '{:,.2f}%',
        'Meta': '{:,.0f}',
        '% Meta vs VN': '{:,.2f}%',
        'Brecha (% S vs M)': '{:+.2f}%'
    }

    row_height = 34
    header_height = 38
    total_rows = len(df_display)
    desired_height = (total_rows * row_height) + header_height + 5

    st.dataframe(
        aplicar_estilo_financiero(df_display).format(formatos),
        use_container_width=True,
        height=desired_height
    )

    if st.button("🚀 Obtener Recomendación IA de Variables Clave", use_container_width=True, type="secondary"):
        with st.spinner("🧠 El CFO Virtual está analizando las desviaciones para darte recomendaciones..."):
            recomendacion_ia_content = generar_recomendacion_variables_ia(df_completo)
            st.session_state['recomendacion_ia_dashboard'] = recomendacion_ia_content
            st.session_state['recomendacion_ia_dashboard_content'] = recomendacion_ia_content

    if 'recomendacion_ia_dashboard' in st.session_state:
        st.markdown("### 💡 Recomendación de Variables Clave por IA")
        st.markdown(st.session_state['recomendacion_ia_dashboard'])
        st.markdown("---")

    if not df_variables_mod.empty:
        st.subheader("🎯 Variables Modificadas")
        st.dataframe(
            df_variables_mod,
            use_container_width=True,
            hide_index=True
        )
        st.info(
            f"**{len(df_variables_mod)}** variables modificadas\n\nEstas son las palancas que estás utilizando para optimizar el resultado."
        )
    else:
        st.info(
            "🔍 **No hay variables modificadas.** Usa los controles del panel lateral para simular diferentes escenarios."
        )

with tab2:
    # ... (Tu código para el análisis visual va aquí)
    # ... (No se necesitan cambios en esta sección, se omite por brevedad)
    st.header("Análisis Visual Intuitivo")
    st.info("Estos gráficos te ayudan a identificar rápidamente los puntos clave de tu simulación.")
    col1_chart, col2_chart = st.columns(2)
    with col1_chart:
        st.subheader("Puente de Utilidad (Waterfall)")
        df_sim = df_completo.set_index('Cuenta')['Simulado']
        fig_waterfall = go.Figure(go.Waterfall(
            name="Simulado", orientation="v",
            measure=['absolute', 'relative', 'total', 'relative', 'total', 'relative', 'total', 'relative', 'absolute'],
            x=['Ventas Netas', 'Costo', 'Margen Bruto', 'Gastos Op.', 'EBITDA Op.', 'Otros Gastos', 'EBITDA',
               'Financieros', 'BAI'],
            y=[df_sim.get('VENTAS NETAS', 0), -df_sim.get('COSTO', 0), 0, -df_sim.get('TOTAL GASTOS OPERATIVOS', 0), 0,
               -df_sim.get('TOTAL DE OTROS GASTOS', 0), 0, -df_sim.get('FINANCIEROS', 0), 0],
            totals={'marker': {'color': PALETA_GRAFICOS['Actual']}},
            increasing={'marker': {'color': PALETA_GRAFICOS['Positivo']}},
            decreasing={'marker': {'color': PALETA_GRAFICOS['Negativo']}}))

        fig_waterfall.update_layout(
            margin=dict(l=20, r=20, t=30, b=20),
            height=400,
            xaxis_tickangle=-45
        )
        st.plotly_chart(fig_waterfall, use_container_width=True, key="waterfall_chart")

    with col2_chart:
        st.subheader("Composición de Costos y Gastos")
        df_costs = df_completo.set_index('Cuenta')
        fig_pie = px.pie(
            values=[
                abs(df_costs.loc['COSTO', 'Simulado']),
                abs(df_costs.loc['TOTAL GASTOS OPERATIVOS', 'Simulado']),
                abs(df_costs.loc['TOTAL DE OTROS GASTOS', 'Simulado']),
                abs(df_costs.loc['FINANCIEROS', 'Simulado'])
            ],
            names=['Costos', 'Gastos Operativos', 'Otros Gastos', 'Financieros'],
            hole=0.4,
            color_discrete_map={
                'Costos': PALETA_GRAFICOS['Egresos'],
                'Gastos Operativos': PALETA_GRAFICOS['Egresos'],
                'Otros Gastos': PALETA_GRAFICOS['Egresos'],
                'Financieros': PALETA_GRAFICOS['Egresos']
            }
        )
        fig_pie.update_layout(
            margin=dict(l=20, r=20, t=30, b=20),
            height=400,
            legend_orientation="h",
            legend_yanchor="bottom",
            legend_y=-0.2
        )
        st.plotly_chart(fig_pie, use_container_width=True, key="pie_chart")

    st.markdown("---")
    st.header("🤖 IA: Análisis Estratégico (Actual vs. Meta vs. Simulación)")
    st.info(
        "Esta herramienta utiliza IA para analizar tus escenarios, proporcionando un análisis profundo y recomendaciones estratégicas, incluyendo razones financieras clave."
    )
    if st.button("🚀 Generar Análisis Estratégico", use_container_width=True, type="primary"):
        with st.spinner("🧠 El CFO Virtual está analizando tus escenarios..."):
            informe = generar_insight_financiero(df_completo)
            st.session_state['informe_ia'] = informe

    if 'informe_ia' in st.session_state:
        st.markdown(st.session_state['informe_ia'])
